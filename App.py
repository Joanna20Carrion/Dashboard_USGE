import streamlit as st
import pandas as pd
import urllib.parse
import requests
import numpy as np
import plotly.express as px
import re
from io import BytesIO
from openpyxl import load_workbook
import unicodedata

st.set_page_config(page_title="Especialistas_Tareas", layout="wide")
st.title("SISTEMAS DE PROCESOS USGE")

tab_tareas, tab_procedimientos, tab_supervisores, tab_empresas_procedimientos, tab_SUP2500037, tab_SUP2400128, tab_SUP2400205, tab_SUP2500029, tab_SUP2400028 = st.tabs(["Tareas", "Procedimientos","Supervisores","Empresas","SUP2500037","SUP2400128","SUP2400205","SUP2500029","SUP2400028"])

with tab_tareas:
    # =========================
    # CONFIG
    # =========================
    SHEET_ID = "1S5z7fRkCuhBgRc-XhMye87xnjnHUtDaK1hf-L-wtWGA"
    SHEET_NAME = "Especialistas_Tareas"
    APPS_SCRIPT_URL = "https://script.google.com/macros/s/AKfycbyXQDjHq6rUVkOAiZre1tWOBC5nWz4SnUjuqKI3xsn2P599vi1ab5z4_57v9qDknk_IfA/exec"

    # =========================
    # LEER GOOGLE SHEET (CSV público)
    # =========================
    sheet = urllib.parse.quote(SHEET_NAME)
    url = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/gviz/tq?tqx=out:csv&sheet={sheet}"

    df = pd.read_csv(url, dtype=str)
    df.columns = (
        df.columns.astype(str)
        .str.replace("\ufeff", "", regex=False)
        .str.strip()
        .str.upper()
    )
    df = df.loc[:, ~df.columns.str.startswith("UNNAMED")]

    # Limpieza mínima
    df["ESPECIALISTA"] = df["ESPECIALISTA"].astype(str).str.strip()
    df["ID_TAREA"] = df["ID_TAREA"].astype(str).str.strip()
    df["ESTADO"] = df["ESTADO"].astype(str).str.strip()
    df = df[(df["ESPECIALISTA"] != "") & (df["ID_TAREA"] != "")]

    # =========================
    # FILTRO POR ESPECIALISTA
    # =========================
    st.header("TAREAS DE ESPECIALISTAS USGE")
    st.subheader("FILTRO")

    lista_especialistas = sorted(df["ESPECIALISTA"].unique().tolist())
    filtro_especialista = st.selectbox(
        "ESPECIALISTA",
        ["TODOS"] + lista_especialistas
    )

    if filtro_especialista != "TODOS":
        df_f = df[df["ESPECIALISTA"] == filtro_especialista].copy()
    else:
        df_f = df.copy()

    # =========================
    # FUNCIÓN PARA SALTO DE LÍNEA EN NOMBRES
    # =========================
    def wrap_name(name, max_chars=18):
        parts = str(name).split()
        if len(parts) <= 2:
            return name
    
        # Construir primera línea hasta el límite
        line1 = []
        for p in parts:
            if len(" ".join(line1 + [p])) <= max_chars:
                line1.append(p)
            else:
                break
    
        line2 = parts[len(line1):]
    
        # Rebalanceo para no dejar apellidos partidos o líneas ridículas
        if len(" ".join(line2)) < 8 and len(line1) > 1:
            line2.insert(0, line1.pop())
    
        return " ".join(line1) + "<br>" + " ".join(line2)
    
    # =========================
    # GRÁFICOS INTERACTIVOS
    # =========================
    st.subheader("RESUMEN")
    
    colA, colB, colC = st.columns(3)
    
    with colA:
        st.metric("Total de tareas", len(df_f))
    
    with colB:
        st.metric("Cumplidas", (df_f["ESTADO"].str.lower() == "cumplido").sum())
    
    with colC:
        st.metric("Pendientes", (df_f["ESTADO"].str.lower() != "cumplido").sum())
        
    col1, col2 = st.columns([2, 1])

    with col1:
        conteo_df = (
            df_f.groupby("ESPECIALISTA")
            .size()
            .reset_index(name="CANT_TAREAS")
            .sort_values("ESPECIALISTA")
        )
        conteo_df["ESPECIALISTA_WRAP"] = conteo_df["ESPECIALISTA"].apply(wrap_name)

        fig_bar = px.bar(
            conteo_df,
            x="ESPECIALISTA_WRAP",
            y="CANT_TAREAS",
            hover_data={"CANT_TAREAS": True},
            labels={"ESPECIALISTA_WRAP": "Especialista", "CANT_TAREAS": "Cantidad de tareas"},
            title="Tareas por especialista"
        )
        fig_bar.update_layout(xaxis_tickangle=0, xaxis_title="Especialista", yaxis_title="Cantidad de tareas")
        st.plotly_chart(fig_bar, width="stretch")

    with col2:
        estado_norm = df_f["ESTADO"].astype(str).str.strip().str.lower()
        resumen = pd.DataFrame({
            "Estado": ["Cumplidas", "Pendientes"],
            "Cantidad": [(estado_norm == "cumplido").sum(), (estado_norm != "cumplido").sum()]
        })

        fig_pie = px.pie(
            resumen,
            names="Estado",
            values="Cantidad",
            hover_data=["Cantidad"],
            title="Cumplidas vs Pendientes"
        )
        st.plotly_chart(fig_pie, width="stretch")

    # =========================
    # TABLA EDITABLE
    # =========================
    st.divider()
    st.subheader("TABLA EDITABLE")

    editable_cols = ["ESPECIALISTA", "EST", "NOMBRE DEL PUESTO", "PRINCIPALES TEMAS ASIGNADOS", "ESTADO"]
    view = df_f[["ID_TAREA"] + editable_cols].copy()

    orig_key = f"orig_{filtro_especialista}"
    if orig_key not in st.session_state:
        st.session_state[orig_key] = view.copy()

    def _norm(x):
        if x is None:
            return ""
        s = str(x)
        return "" if s.lower() == "nan" else s.strip()

    edited = st.data_editor(
        view,
        width="stretch",
        hide_index=True,
        num_rows="fixed",
        key=f"editor_{filtro_especialista}",
    )

    # =========================
    # GUARDAR CAMBIOS EN LOTE
    # =========================
    if st.button("💾 Guardar Tarea"):
        original = st.session_state[orig_key].copy()

        changed = pd.Series(np.zeros(len(edited), dtype=bool), index=edited.index)
        for c in editable_cols:
            changed |= edited[c].map(_norm) != original[c].map(_norm)

        changed_rows = edited.loc[changed].copy()

        if changed_rows.empty:
            st.info("No hay cambios.")
        else:
            updates = []
            for _, r in changed_rows.iterrows():
                updates.append({
                    "ID_TAREA": _norm(r["ID_TAREA"]),
                    "ESPECIALISTA": _norm(r["ESPECIALISTA"]),
                    "EST": _norm(r["EST"]),
                    "NOMBRE DEL PUESTO": _norm(r["NOMBRE DEL PUESTO"]),
                    "PRINCIPALES TEMAS ASIGNADOS": _norm(r["PRINCIPALES TEMAS ASIGNADOS"]),
                    "ESTADO": _norm(r["ESTADO"]),
                })

            payload = {"action": "BATCH_UPDATE", "updates": updates}
            resp = requests.post(APPS_SCRIPT_URL, json=payload, timeout=30)
            data = resp.json() if resp.ok else {"ok": False, "error": resp.text}

            if data.get("ok"):
                st.success(f"Guardado ✅ Filas actualizadas: {data.get('updated')}")
                st.rerun()
            else:
                st.error(data)

    # =========================
    # AGREGAR FILA
    # =========================
    st.divider()
    st.subheader("AGREGAR FILA")

    with st.form("add_row"):
        especialista = st.selectbox(
            "ESPECIALISTA",
            [
                "EDUARDO CARRILLO TINCALLPA",
                "JORGE ISRAEL MONTENEGRO SANTOS",
                "RUBEN ROJAS RAMIREZ",
                "JORGE PEDRO VILCACHAGUA NUÑEZ",
                "CESAR GUILLERMO OLANO OCHOA",
                "ANGEL DANIEL ROBLES SARAVIA",
                "GERMAN ABEL GUTARRA CRIBILLERO",
            ]
        )
        
        est = st.selectbox("EST", [f"EST {i}" for i in range(1, 8)])
        puesto = st.text_input("NOMBRE DEL PUESTO")
        temas = st.text_area("PRINCIPALES TEMAS ASIGNADOS")
        estado = st.selectbox("ESTADO", ["Pendiente", "Cumplido"])

        add_ok = st.form_submit_button("➕ Agregar")

    if add_ok:
        payload = {
            "action": "ADD",
            "especialista": especialista.strip(),
            "est": est.strip(),
            "puesto": puesto.strip(),
            "temas": temas.strip(),
            "estado": estado,
        }
        resp = requests.post(APPS_SCRIPT_URL, json=payload, timeout=30)
        data = resp.json() if resp.ok else {"ok": False, "error": resp.text}

        if data.get("ok"):
            st.success(f"Agregado ✅ ID_TAREA = {data.get('id')}")
            st.rerun()
        else:
            st.error(data)

    # =========================
    # BORRAR FILA
    # =========================
    st.divider()
    st.subheader("BORRAR FILA")

    id_del = st.selectbox("ID_TAREA a borrar", df["ID_TAREA"].astype(str).tolist())
    confirm = st.checkbox("Confirmo borrado irreversible")

    if st.button("🗑️ Borrar"):
        if not confirm:
            st.warning("Confirma antes de borrar.")
        else:
            payload = {"action": "DELETE", "id": str(id_del)}
            resp = requests.post(APPS_SCRIPT_URL, json=payload, timeout=30)
            data = resp.json() if resp.ok else {"ok": False, "error": resp.text}

            if data.get("ok"):
                st.success("Borrado ✅")
                st.rerun()
            else:
                st.error(data)

with tab_procedimientos:
    st.header("PROCEDIMIENTOS USGE")

    SHEET_NAME_P = "Acciones_Procedimientos"

    # -------- Leer Google Sheet (CSV público) --------
    sheet_p = urllib.parse.quote(SHEET_NAME_P)
    url_p = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/gviz/tq?tqx=out:csv&sheet={sheet_p}"

    dfp = pd.read_csv(url_p, dtype=str)
    dfp.columns = (
        dfp.columns.astype(str)
        .str.replace("\ufeff", "", regex=False)
        .str.strip()
        .str.upper()
    )
    dfp = dfp.loc[:, ~dfp.columns.str.startswith("UNNAMED")]

    # Clave real
    if "ID_PR" not in dfp.columns:
        st.error("La hoja 'Acciones_Procedimientos' debe tener una columna llamada 'ID_PR' (clave única).")
        st.stop()

    # Limpieza mínima
    dfp["ID_PR"] = dfp["ID_PR"].astype(str).str.strip()
    dfp = dfp[dfp["ID_PR"] != ""]

    # =========================
    # FILTROS (4)
    # =========================
    st.subheader("FILTRO")

    colf1, colf2, colf3, colf4 = st.columns(4)

    opts_esp = ["TODOS"] + sorted(dfp["ESPECIALISTA"].dropna().astype(str).str.strip().unique().tolist())
    opts_plan = ["TODOS"] + sorted(dfp["PLAN DE ACCIÓN ESPECÍFICO"].dropna().astype(str).str.strip().unique().tolist())
    opts_acc = ["TODOS"] + sorted(dfp["ACCIONES FÍSICAS"].dropna().astype(str).str.strip().unique().tolist())
    opts_proc = ["TODOS"] + sorted(dfp["PROCEDIMIENTO"].dropna().astype(str).str.strip().unique().tolist())

    with colf1:
        f_esp = st.selectbox("ESPECIALISTA", opts_esp, key="f_proc_esp")

    with colf2:
        f_plan = st.selectbox("PLAN DE ACCIÓN", opts_plan, key="f_proc_plan")

    with colf3:
        f_acc = st.selectbox("ACCIONES FÍSICAS", opts_acc, key="f_proc_acc")

    with colf4:
        f_pro = st.selectbox("PROCEDIMIENTO", opts_proc, key="f_proc_pro")

    # Aplicar filtros
    dfp_f = dfp.copy()

    if f_esp != "TODOS":
        dfp_f = dfp_f[dfp_f["ESPECIALISTA"].astype(str).str.strip() == f_esp]

    if f_plan != "TODOS":
        dfp_f = dfp_f[dfp_f["PLAN DE ACCIÓN ESPECÍFICO"].astype(str).str.strip() == f_plan]

    if f_acc != "TODOS":
        dfp_f = dfp_f[dfp_f["ACCIONES FÍSICAS"].astype(str).str.strip() == f_acc]

    if f_pro != "TODOS":
        dfp_f = dfp_f[dfp_f["PROCEDIMIENTO"].astype(str).str.strip() == f_pro]

    # =========================
    # RESUMEN + GRÁFICOS
    # =========================
    st.subheader("RESUMEN")

    # --- Cargar hoja de TAREAS para traer EST por ESPECIALISTA ---
    SHEET_NAME_T = "Especialistas_Tareas"
    sheet_t = urllib.parse.quote(SHEET_NAME_T)
    url_t = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/gviz/tq?tqx=out:csv&sheet={sheet_t}"

    dft = pd.read_csv(url_t, dtype=str)
    dft.columns = (
        dft.columns.astype(str)
        .str.replace("\ufeff", "", regex=False)
        .str.strip()
        .str.upper()
    )
    dft = dft.loc[:, ~dft.columns.str.startswith("UNNAMED")]

    dft["ESPECIALISTA"] = dft["ESPECIALISTA"].astype(str).str.strip()
    dft["EST"] = dft["EST"].astype(str).str.strip()

    est_map = (
        dft[dft["EST"].notna() & (dft["EST"] != "")]
        .drop_duplicates(subset=["ESPECIALISTA"])
        .set_index("ESPECIALISTA")["EST"]
        .to_dict()
    )

    # --- KPIs ---
    # total de PR únicos (sin repetir)
    total_pr_unicos = (
        dfp_f["PROCEDIMIENTO"].dropna().astype(str).str.strip().replace("", np.nan).dropna().nunique()
        if "PROCEDIMIENTO" in dfp_f.columns else 0
    )
    uniq_esp = dfp_f["ESPECIALISTA"].dropna().astype(str).str.strip().replace("", np.nan).dropna().nunique()

    k1, k2 = st.columns(2)
    k1.metric("Procedimientos", total_pr_unicos)
    k2.metric("Especialistas", uniq_esp)

    # --- Tabla resumen por especialista: EST + PR únicos (lista) + conteos ---
    if not dfp_f.empty:
        dfp_tmp = dfp_f.copy()
        dfp_tmp["ESPECIALISTA"] = dfp_tmp["ESPECIALISTA"].astype(str).str.strip()
        dfp_tmp["PROCEDIMIENTO"] = dfp_tmp["PROCEDIMIENTO"].astype(str).str.strip()
        dfp_tmp["ACCIONES FÍSICAS"] = dfp_tmp["ACCIONES FÍSICAS"].astype(str).str.strip()
    
        def lista_unicos(series: pd.Series) -> str:
            vals = (
                series.replace("", np.nan)
                .dropna()
                .astype(str)
                .str.strip()
                .unique()
                .tolist()
            )
            vals = sorted([v for v in vals if v])
            return "\n".join(vals) if vals else "—"
    
        resumen_esp = (
            dfp_tmp
            .groupby("ESPECIALISTA", dropna=False)
            .agg(
                N_PR_UNICOS=("PROCEDIMIENTO", lambda x: x.replace("", np.nan).dropna().nunique()),
                N_ACCIONES_FISICAS_UNICAS=("ACCIONES FÍSICAS", lambda x: x.replace("", np.nan).dropna().nunique()),
            )
            .reset_index()
        )
        
        resumen_esp["EST"] = resumen_esp["ESPECIALISTA"].map(est_map).fillna("—")
        
        st.markdown("**Resumen por especialista**")
        
        resumen_esp = resumen_esp.rename(columns={
            "ESPECIALISTA": "Especialista",
            "EST": "EST",
            "N_PR_UNICOS": "N° Procedimientos",
            "PR_UNICOS": "Procedimientos",
            "N_ACCIONES_FISICAS_UNICAS": "N° Acciones físicas",
        })

        st.dataframe(resumen_esp, width="stretch", hide_index=True)

    # --- wrap eje X en 2 líneas ---
    def wrap_x(name, max_chars=18):
        parts = str(name).split()
        if len(parts) <= 2:
            return name
    
        # construir la primera línea hasta el límite
        line1 = []
        for p in parts:
            if len(" ".join(line1 + [p])) <= max_chars:
                line1.append(p)
            else:
                break
    
        line2 = parts[len(line1):]
    
        # si la segunda línea queda muy corta, redistribuir
        if len(" ".join(line2)) < 8 and len(line1) > 1:
            line2.insert(0, line1.pop())
    
        return " ".join(line1) + "<br>" + " ".join(line2)
    
    # =========================
    # GRÁFICOS
    # =========================
    cbar, cpie = st.columns([2, 1])

    # Barras: PR únicos por especialista (sin repetir)
    with cbar:
        conteo_proc = (
            dfp_f[["ESPECIALISTA", "PROCEDIMIENTO"]]
            .dropna()
            .assign(
                ESPECIALISTA=lambda x: x["ESPECIALISTA"].astype(str).str.strip(),
                PROCEDIMIENTO=lambda x: x["PROCEDIMIENTO"].astype(str).str.strip(),
            )
            .query("ESPECIALISTA != '' and PROCEDIMIENTO != ''")
            .groupby("ESPECIALISTA")["PROCEDIMIENTO"]
            .nunique()
            .reset_index(name="N_PR_UNICOS")
            .sort_values("N_PR_UNICOS", ascending=False)
        )

        conteo_proc["ESPECIALISTA_WRAP"] = conteo_proc["ESPECIALISTA"].apply(wrap_x)

        fig_bar = px.bar(
            conteo_proc,
            x="ESPECIALISTA_WRAP",
            y="N_PR_UNICOS",
            hover_data=["N_PR_UNICOS"],
            labels={
                "ESPECIALISTA_WRAP": "Especialista",
                "N_PR_UNICOS": "PR únicos"
            },
            title="Procedimientos por especialista"
        )

        fig_bar.update_layout(
            xaxis_tickangle=0,
            xaxis_title="Especialista",
            yaxis_title="Cantidad de procedimientos"
        )

        st.plotly_chart(fig_bar, width="stretch")

    # Circular: Acciones físicas únicas por especialista
    with cpie:
        acciones_por_esp = (
            dfp_f[["ESPECIALISTA", "ACCIONES FÍSICAS"]]
            .dropna()
            .assign(
                ESPECIALISTA=lambda x: x["ESPECIALISTA"].astype(str).str.strip(),
                ACCIONES_FISICAS=lambda x: x["ACCIONES FÍSICAS"].astype(str).str.strip(),
            )
            .query("ESPECIALISTA != '' and ACCIONES_FISICAS != ''")
            .groupby("ESPECIALISTA")["ACCIONES_FISICAS"]
            .nunique()
            .reset_index(name="CANT_ACCIONES_FISICAS_UNICAS")
            .sort_values("CANT_ACCIONES_FISICAS_UNICAS", ascending=False)
        )

        fig_pie_p = px.pie(
            acciones_por_esp,
            names="ESPECIALISTA",
            values="CANT_ACCIONES_FISICAS_UNICAS",
            hover_data=["CANT_ACCIONES_FISICAS_UNICAS"],
            title="Acciones físicas por especialista"
        )

        st.plotly_chart(fig_pie_p, width="stretch")

    # =========================
    # TABLA EDITABLE
    # =========================
    st.divider()
    st.subheader("TABLA EDITABLE")

    editable_cols_p = [
        "PLAN DE ACCIÓN ESPECÍFICO",
        "ACCIONES FÍSICAS",
        "RCD/RD/PR/LEY",
        "PROCEDIMIENTO",
        "ESPECIALISTA",
        "QUÉ INFORMACIÓN SE REQUIERE EN CADA PR",
        "QUÉ INFORMACIÓN SE RECOPILA EN CADA PR",
    ]

    missing = [c for c in editable_cols_p if c not in dfp.columns]
    if missing:
        st.error(f"Faltan columnas en la hoja: {missing}")
        st.stop()

    viewp = dfp_f[["ID_PR"] + editable_cols_p].copy()

    orig_key_p = f"orig_proc_{f_esp}_{f_plan}_{f_acc}_{f_pro}"
    if orig_key_p not in st.session_state:
        st.session_state[orig_key_p] = viewp.copy()

    def _norm_p(x):
        if x is None:
            return ""
        s = str(x)
        return "" if s.lower() == "nan" else s.strip()

    editedp = st.data_editor(
        viewp,
        width="stretch",
        hide_index=True,
        num_rows="fixed",
        key=f"editor_proc_{f_esp}_{f_plan}_{f_acc}_{f_pro}",
    )

    # =========================
    # GUARDAR CAMBIOS
    # =========================
    if st.button("💾 Guardar Procedimiento"):
        original = st.session_state[orig_key_p].copy()

        changed = pd.Series(np.zeros(len(editedp), dtype=bool), index=editedp.index)
        for c in editable_cols_p:
            changed |= editedp[c].map(_norm_p) != original[c].map(_norm_p)

        changed_rows = editedp.loc[changed].copy()

        if changed_rows.empty:
            st.info("No hay cambios.")
        else:
            updates = []
            for _, r in changed_rows.iterrows():
                fields = {c: _norm_p(r[c]) for c in editable_cols_p}
                updates.append({
                    "ID": _norm_p(r["ID_PR"]),  # Apps Script espera "ID"
                    "FIELDS": fields
                })

            payload = {
                "action": "BATCH_UPDATE",
                "sheet": SHEET_NAME_P,
                "updates": updates
            }

            resp = requests.post(APPS_SCRIPT_URL, json=payload, timeout=30)
            data = resp.json() if resp.ok else {"ok": False, "error": resp.text}

            if data.get("ok"):
                st.success(f"Guardado ✅ Actualizadas: {data.get('updated')} | No encontradas: {data.get('notFound')}")
                st.session_state[orig_key_p] = editedp.copy()
                st.rerun()
            else:
                st.error(data)

    # =========================
    # AGREGAR FILA
    # =========================
    st.divider()
    st.subheader("AGREGAR PROCEDIMIENTO")

    with st.form("add_row_procedimientos"):
        plan = st.selectbox(
            "PLAN DE ACCIÓN ESPECÍFICO",
            [
                "1.1 Supervisión de la Generación del SEIN",
                "1.2 Supervisión del COES",
                "1.3 Supervisión de la Regulación del Mercado Eléctrico en los Agentes SEIN_COES",
                "1.4 Supervisión de la Operatividad de la Generación en Sistemas Eléctricos Aislados y la Generación Distribuida a cargo de las Empresas Distribuidoras y Autoproductoras",
                "1.5 Supervisión de las condiciones de la Generación Eléctrica para la prestación del servicio público de electricidad en los sistemas aislados y la Operatividad de los Sistemas Fotovoltaicos no conectados a Red",
                "1.6 Supervisión del Cumplimiento de la Norma Técnica de Calidad de los Servicios Eléctricos (NTCSE) y su base metodológicas",
            ]
        )

        acciones = st.selectbox(
            "ACCIONES FÍSICAS",
            [
                "Supervisión de la disponibilidad y operatividad unidades SEIN",
                "Supervisión ensayos unidades generación",
                "Supervisión del desempeño de las unidades de generación despachada por el COES",
                "Supervisión de las inflexibilidades operativas de las unidades de generación",
                "Supervisión programas de mantenimiento COES",
                "Supervisión de la programación de la operación",
                "Supervisión de coordinación de operación",
                "Supervisión evaluación de operación",
                "Supervisión esquemas rechazo carga generación",
                "Monitoreo operación en tiempo real",
                "Operación NETMECOTR y la NTIITR",
                "Operación semanal y margen de reserva",
                "Evaluación de procedimientos del COES",
                "Supervisión recaudación FISE - Generación",
                "Seguimiento de contratos de usuarios libres",
                "Supervisión costos incurridos de generación adicional",
                "Supervisión transacciones económicas",
                "Supervisión de generación aislada",
                "Supervisión de generación no despachada por el COES",
                "Supervisión de autoproductoras",
                "Supervisión de accidentes de terceros",
                "Supervisión de sistemas fotovoltaicos autónomos",
                "Supervisión de la Norma Técnica de calidad de los servicios eléctricos y su base metodológica",
            ]
        )

        rcd = st.text_input("RCD/RD/PR/LEY")
        procedimiento = st.text_area("PROCEDIMIENTO")

        especialista = st.selectbox(
            "ESPECIALISTA",
            [
                "EDUARDO CARRILLO TINCALLPA",
                "JORGE ISRAEL MONTENEGRO SANTOS",
                "RUBEN ROJAS RAMIREZ",
                "JORGE PEDRO VILCACHAGUA NUÑEZ",
                "CESAR GUILLERMO OLANO OCHOA",
                "ANGEL DANIEL ROBLES SARAVIA",
                "GERMAN ABEL GUTARRA CRIBILLERO",
            ]
        )

        info_req = st.text_area("¿QUÉ INFORMACIÓN SE REQUIERE EN EL PROCEDIMIENTO?")
        info_rec = st.text_area("¿QUÉ INFORMACIÓN SE RECOPILA EN EL PROCEDIMIENTO?")

        add_ok_p = st.form_submit_button("➕ Agregar")

    if add_ok_p:
        payload = {
            "action": "ADD",
            "sheet": SHEET_NAME_P,
            "row": {
                "PLAN DE ACCIÓN ESPECÍFICO": plan,
                "ACCIONES FÍSICAS": acciones,
                "RCD/RD/PR/LEY": rcd,
                "PROCEDIMIENTO": procedimiento,
                "ESPECIALISTA": especialista,
                "QUÉ INFORMACIÓN SE REQUIERE EN CADA PR": info_req,
                "QUÉ INFORMACIÓN SE RECOPILA EN CADA PR": info_rec,
            }
        }

        resp = requests.post(APPS_SCRIPT_URL, json=payload, timeout=30)
        data = resp.json() if resp.ok else {"ok": False, "error": resp.text}

        if data.get("ok"):
            st.success(f"Agregado ✅ ID_PR = {data.get('id')}")
            st.rerun()
        else:
            st.error(data)

    # =========================
    # BORRAR FILA
    # =========================
    st.divider()
    st.subheader("BORRAR PROCEDIMIENTO")

    ids_borrado = dfp_f["ID_PR"].astype(str).tolist() if not dfp_f.empty else dfp["ID_PR"].astype(str).tolist()
    id_del_p = st.selectbox("ID_PR a borrar", ids_borrado, key="del_proc")

    confirm_p = st.checkbox("Confirmo borrado irreversible", key="conf_proc")

    if st.button("🗑️ Borrar Procedimiento", key="btn_del_proc"):
        if not confirm_p:
            st.warning("Confirma antes de borrar.")
        else:
            payload = {"action": "DELETE", "sheet": SHEET_NAME_P, "id": str(id_del_p)}
            resp = requests.post(APPS_SCRIPT_URL, json=payload, timeout=30)
            data = resp.json() if resp.ok else {"ok": False, "error": resp.text}

            if data.get("ok"):
                st.success("Borrado ✅")
                st.rerun()
            else:
                st.error(data)

# =========================
# TAB: SUPERVISORES
# =========================
with tab_supervisores:
    st.header("SUPERVISORES USGE")

    SHEET_NAME_S = "Supervisores"

    # ---- leer CSV público ----
    sheet_s = urllib.parse.quote(SHEET_NAME_S)
    url_s = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/gviz/tq?tqx=out:csv&sheet={sheet_s}"

    dfs = pd.read_csv(url_s, dtype=str)
    dfs.columns = (
        dfs.columns.astype(str)
        .str.replace("\ufeff", "", regex=False)
        .str.strip()
        .str.upper()
    )
    dfs = dfs.loc[:, ~dfs.columns.str.startswith("UNNAMED")]

    # ✅ clave única
    if "ID_SUP" not in dfs.columns:
        st.error("La hoja 'Supervisores' debe tener una columna 'ID_SUP' (clave única).")
        st.stop()

    # Limpieza mínima
    dfs["ID_SUP"] = dfs["ID_SUP"].astype(str).str.strip()
    dfs = dfs[dfs["ID_SUP"] != ""]

    # =========================
    # FILTRO
    # =========================
    st.subheader("FILTRO")

    def _clean_series(s):
        return (
            s.dropna()
            .astype(str)
            .str.strip()
            .replace("", np.nan)
            .dropna()
        )

    # Opciones (dinámicas desde la data)
    opts_especialista = ["TODOS"] + sorted(_clean_series(dfs["ESPECIALISTA"]).unique().tolist())
    opts_est = ["TODOS"] + sorted(_clean_series(dfs["EST"]).unique().tolist())
    opts_supervisor = ["TODOS"] + sorted(_clean_series(dfs["NOMBRE COMPLETO DE SUPERVISOR"]).unique().tolist())
    opts_categoria = ["TODOS"] + sorted(_clean_series(dfs["CATEGORÍA"]).unique().tolist())
    opts_contrato = ["TODOS"] + sorted(_clean_series(dfs["CONTRATO"]).unique().tolist())
    opts_area = ["TODOS"] + sorted(_clean_series(dfs["ÁREA"]).unique().tolist())
    
    c1, c2, c3, c4, c5, c6 = st.columns(6)

    with c1:
        f_esp = st.selectbox("ESPECIALISTA", opts_especialista, key="f_sup_esp")
    with c2:
        f_est = st.selectbox("EST", opts_est, key="f_sup_est")
    with c3:
        f_sup = st.selectbox("SUPERVISOR", opts_supervisor, key="f_sup_sup")
    with c4:
        f_cat = st.selectbox("CATEGORÍA", opts_categoria, key="f_sup_cat")
    with c5:
        f_con = st.selectbox("CONTRATO", opts_contrato, key="f_sup_con")
    with c6:
        f_area = st.selectbox("ÁREA", opts_area, key="f_sup_area")
        
    # Aplicar filtros
    dfs_f = dfs.copy()

    if f_esp != "TODOS":
        dfs_f = dfs_f[dfs_f["ESPECIALISTA"].astype(str).str.strip() == f_esp]
    if f_est != "TODOS":
        dfs_f = dfs_f[dfs_f["EST"].astype(str).str.strip() == f_est]
    if f_sup != "TODOS":
        dfs_f = dfs_f[dfs_f["NOMBRE COMPLETO DE SUPERVISOR"].astype(str).str.strip() == f_sup]
    if f_cat != "TODOS":
        dfs_f = dfs_f[dfs_f["CATEGORÍA"].astype(str).str.strip() == f_cat]
    if f_con != "TODOS":
        dfs_f = dfs_f[dfs_f["CONTRATO"].astype(str).str.strip() == f_con]
    if f_area != "TODOS":
        dfs_f = dfs_f[dfs_f["ÁREA"].astype(str).str.strip() == f_area]
        
    # =========================
    # RESUMEN
    # =========================
    st.subheader("RESUMEN")

    # Fechas: INICIO / FINAL (si no parsea, igual muestra valores)
    ini_vals = _clean_series(dfs_f["INICIO"]) if "INICIO" in dfs_f.columns else pd.Series([], dtype=str)
    fin_vals = _clean_series(dfs_f["FINAL"]) if "FINAL" in dfs_f.columns else pd.Series([], dtype=str)

    # intentar convertir a fecha para min/max
    ini_dt = pd.to_datetime(ini_vals, errors="coerce", dayfirst=True)
    fin_dt = pd.to_datetime(fin_vals, errors="coerce", dayfirst=True)

    fecha_inicio = ini_dt.min()
    fecha_final = fin_dt.max()

    # Cantidad de supervisores (únicos por ID_SUP)
    cant_sup = dfs_f["ID_SUP"].nunique()

    r1, r2, r3 = st.columns(3)

    with r1:
        st.metric("Fecha de Inicio", "—" if pd.isna(fecha_inicio) else fecha_inicio.date().isoformat())
    with r2:
        st.metric("Fecha Final", "—" if pd.isna(fecha_final) else fecha_final.date().isoformat())
    with r3:
        st.metric("Cantidad de Supervisores", int(cant_sup))
    
    # =========================
    # TABLA RESUMEN + GRÁFICOS (antes de TABLA EDITABLE)
    # =========================
    st.markdown("**Resumen por supervisor**")
    
    # Columnas solicitadas (con nombres reales en tu sheet)
    cols_resumen = [
        "NOMBRE COMPLETO DE SUPERVISOR",
        "CATEGORÍA",
        "DNI",
        "CORREO EMPRESARIAL",
        "CORREO PARTICULAR",
        "TELEFONO",
    ]
    
    missing_res = [c for c in cols_resumen if c not in dfs_f.columns]
    if missing_res:
        st.error(f"Faltan columnas para la tabla resumen: {missing_res}")
    else:
        tabla_resumen = dfs_f[cols_resumen].copy()
    
        # limpiar strings
        for c in cols_resumen:
            tabla_resumen[c] = tabla_resumen[c].astype(str).str.strip()
    
        # quitar filas sin nombre de supervisor (si hubiera)
        tabla_resumen = tabla_resumen[tabla_resumen["NOMBRE COMPLETO DE SUPERVISOR"] != ""]
    
        st.dataframe(tabla_resumen, width="stretch", hide_index=True)
    
    # =========================
    # GRÁFICOS: CATEGORÍA y EST
    # =========================    
    g1, g2 = st.columns([1, 1])
    
    # --- Pie: Supervisores por categoría ---
    with g1:
        if "CATEGORÍA" in dfs_f.columns:
            cat_df = (
                dfs_f[["ID_SUP", "CATEGORÍA"]]
                .dropna()
                .assign(
                    CATEGORIA=lambda x: x["CATEGORÍA"].astype(str).str.strip(),
                    ID_SUP=lambda x: x["ID_SUP"].astype(str).str.strip(),
                )
            )
            cat_df = cat_df[(cat_df["CATEGORIA"] != "") & (cat_df["ID_SUP"] != "")]
    
            pie_cat = (
                cat_df.groupby("CATEGORIA")["ID_SUP"]
                .nunique()  # ✅ supervisores únicos
                .reset_index(name="N_SUPERVISORES")
                .sort_values("N_SUPERVISORES", ascending=False)
            )
    
            fig_pie_cat = px.pie(
                pie_cat,
                names="CATEGORIA",
                values="N_SUPERVISORES",
                hover_data=["N_SUPERVISORES"],
                title="Supervisores por categoría"
            )
            st.plotly_chart(fig_pie_cat, width="stretch")
        else:
            st.info("No existe la columna CATEGORÍA.")
    
    # --- Treemap: Supervisores por EST ---
    with g2:
        if "EST" in dfs_f.columns:
            est_df = (
                dfs_f[["ID_SUP", "EST"]]
                .dropna()
                .assign(
                    EST_=lambda x: x["EST"].astype(str).str.strip(),
                    ID_SUP=lambda x: x["ID_SUP"].astype(str).str.strip(),
                )
            )
            est_df = est_df[(est_df["EST_"] != "") & (est_df["ID_SUP"] != "")]
    
            tree_est = (
                est_df.groupby("EST_")["ID_SUP"]
                .nunique()  # ✅ supervisores únicos
                .reset_index(name="N_SUPERVISORES")
                .sort_values("N_SUPERVISORES", ascending=False)
            )
    
            fig_tree = px.treemap(
                tree_est,
                path=["EST_"],
                values="N_SUPERVISORES",
                hover_data=["N_SUPERVISORES"],
                title="Supervisores por EST"
            )
            st.plotly_chart(fig_tree, width="stretch")
        else:
            st.info("No existe la columna EST.")
            
    # =========================
    # TABLA EDITABLE (usa dfs_f)
    # =========================
    st.divider()
    st.subheader("TABLA EDITABLE")

    editable_cols_s = [
        "EST",
        "NOMBRE DE EST",
        "ESPECIALISTA",
        "CONTRATO",
        "CATEGORÍA",
        "NOMBRE COMPLETO DE SUPERVISOR",
        "DNI",
        "CORREO EMPRESARIAL",
        "CORREO PARTICULAR",
        "TELEFONO",
        "INICIO",
        "FINAL",
        "PERFIL",
        "FECHA DE NACIMIENTO",
        "ÁREA",
    ]

    missing = [c for c in editable_cols_s if c not in dfs.columns]
    if missing:
        st.error(f"Faltan columnas en 'Supervisores': {missing}")
        st.stop()

    views = dfs_f[["ID_SUP"] + editable_cols_s].copy()

    orig_key_s = f"orig_supervisores_{f_esp}_{f_est}_{f_sup}_{f_cat}_{f_con}"
    if orig_key_s not in st.session_state:
        st.session_state[orig_key_s] = views.copy()

    def _norm_s(x):
        if x is None:
            return ""
        s = str(x)
        return "" if s.lower() == "nan" else s.strip()

    editeds = st.data_editor(
        views,
        width="stretch",
        hide_index=True,
        num_rows="fixed",
        key=f"editor_supervisores_{f_esp}_{f_est}_{f_sup}_{f_cat}_{f_con}",
    )

    # =========================
    # GUARDAR CAMBIOS EN LOTE
    # =========================
    if st.button("💾 Guardar Supervisor"):
        original = st.session_state[orig_key_s].copy()

        changed = pd.Series(np.zeros(len(editeds), dtype=bool), index=editeds.index)
        for c in editable_cols_s:
            changed |= editeds[c].map(_norm_s) != original[c].map(_norm_s)

        changed_rows = editeds.loc[changed].copy()

        if changed_rows.empty:
            st.info("No hay cambios.")
        else:
            updates = []
            for _, r in changed_rows.iterrows():
                fields = {c: _norm_s(r[c]) for c in editable_cols_s}
                updates.append({"ID": _norm_s(r["ID_SUP"]), "FIELDS": fields})

            payload = {"action": "BATCH_UPDATE", "sheet": SHEET_NAME_S, "updates": updates}
            resp = requests.post(APPS_SCRIPT_URL, json=payload, timeout=30)
            data = resp.json() if resp.ok else {"ok": False, "error": resp.text}

            if data.get("ok"):
                st.success(f"Guardado ✅ Actualizadas: {data.get('updated')} | No encontradas: {data.get('notFound')}")
                st.session_state[orig_key_s] = editeds.copy()
                st.rerun()
            else:
                st.error(data)

    # =========================
    # AGREGAR FILA (tu bloque actual, igual)
    # =========================
    st.divider()
    st.subheader("AGREGAR SUPERVISOR")

    with st.form("add_row_supervisores"):
        est = st.selectbox("EST", [f"EST {i}" for i in range(1, 7)])

        nombre_est = st.selectbox(
            "NOMBRE DE EST",
            [
                "GREEN ELECTRICITY ENGINEERING CORPORATION SUCURSAL DEL PERU",
                "VISION QUALITY ENERGY S.A.C.",
                "SEOUL INSPECTION",
                "VASMOL S.A.C.",
                "OTRO",
            ],
        )

        especialista = st.selectbox(
            "ESPECIALISTA",
            [
                "EDUARDO CARRILLO TINCALLPA",
                "JORGE ISRAEL MONTENEGRO SANTOS",
                "RUBEN ROJAS RAMIREZ",
                "JORGE PEDRO VILCACHAGUA NUÑEZ",
                "CESAR GUILLERMO OLANO OCHOA",
                "ANGEL DANIEL ROBLES SARAVIA",
                "GERMAN ABEL GUTARRA CRIBILLERO",
            ],
        )

        contrato = st.selectbox(
            "CONTRATO",
            [
                "SUP2400128",
                "SUP2400028 (item 4)",
                "SUP2500037",
                "SUP2500029",
                "SUP2500023",
                "SUP2400178",
                "SUP2400087",
                "SUP2400177",
                "SUP2500133",
                "SUP2500082",
                "SUP2500046",
                "SUP2500074",
                "SUP2500081",
                "SUP2500028",
                "OTRO",
            ],
        )

        categoria = st.selectbox(
            "CATEGORÍA",
            ["S1A", "S1B", "S2", "S3A", "S3B", "S4", "S4A", "S4B"],
        )

        area = st.selectbox("ÁREA", ["GENERACIÓN", "TRANSMISIÓN", "OTRO"])

        nombre_sup = st.text_input("NOMBRE COMPLETO DE SUPERVISOR")
        dni = st.text_input("DNI")
        correo_emp = st.text_input("CORREO EMPRESARIAL")
        correo_part = st.text_input("CORREO PARTICULAR")
        telefono = st.text_input("TELEFONO")
        inicio = st.text_input("FECHA DE INICIO")
        final = st.text_input("FECHA FINAL")
        perfil = st.text_input("PERFIL")
        fnac = st.text_input("FECHA DE NACIMIENTO")

        add_ok_s = st.form_submit_button("➕ Agregar")

    if add_ok_s:
        payload = {
            "action": "ADD",
            "sheet": SHEET_NAME_S,
            "row": {
                "EST": est.strip(),
                "NOMBRE DE EST": nombre_est.strip(),
                "ESPECIALISTA": especialista.strip(),
                "CONTRATO": contrato.strip(),
                "CATEGORÍA": categoria.strip(),
                "NOMBRE COMPLETO DE SUPERVISOR": nombre_sup.strip(),
                "DNI": dni.strip(),
                "CORREO EMPRESARIAL": correo_emp.strip(),
                "CORREO PARTICULAR": correo_part.strip(),
                "TELEFONO": telefono.strip(),
                "INICIO": inicio.strip(),
                "FINAL": final.strip(),
                "PERFIL": perfil.strip(),
                "FECHA DE NACIMIENTO": fnac.strip(),
                "ÁREA": area.strip(),
            },
        }

        resp = requests.post(APPS_SCRIPT_URL, json=payload, timeout=30)
        data = resp.json() if resp.ok else {"ok": False, "error": resp.text}

        if data.get("ok"):
            st.success(f"Agregado ✅ ID_SUP = {data.get('id')}")
            st.rerun()
        else:
            st.error(data)

    # =========================
    # BORRAR FILA (usa dfs_f si tiene filas)
    # =========================
    st.divider()
    st.subheader("BORRAR SUPERVISOR")

    ids_borrado = dfs_f["ID_SUP"].astype(str).tolist() if not dfs_f.empty else dfs["ID_SUP"].astype(str).tolist()
    id_del = st.selectbox("ID_SUP a borrar", ids_borrado, key="del_sup")

    confirm_s = st.checkbox("Confirmo borrado irreversible", key="conf_sup")

    if st.button("🗑️ Borrar Supervisor", key="btn_del_sup"):
        if not confirm_s:
            st.warning("Confirma antes de borrar.")
        else:
            payload = {"action": "DELETE", "sheet": SHEET_NAME_S, "id": str(id_del)}
            resp = requests.post(APPS_SCRIPT_URL, json=payload, timeout=30)
            data = resp.json() if resp.ok else {"ok": False, "error": resp.text}

            if data.get("ok"):
                st.success("Borrado ✅")
                st.rerun()
            else:
                st.error(data)

# =========================
# TAB: EMPRESAS - PROCEDIMIENTOS
# =========================
with tab_empresas_procedimientos:
    st.header("EMPRESAS DE PROCEDIMIENTOS USGE")

    SHEET_NAME_EP = "Empresas_Procedimientos"

    # ---- leer CSV público ----
    sheet_ep = urllib.parse.quote(SHEET_NAME_EP)
    url_ep = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/gviz/tq?tqx=out:csv&sheet={sheet_ep}"

    dfep = pd.read_csv(url_ep, dtype=str)
    dfep.columns = (
        dfep.columns.astype(str)
        .str.replace("\ufeff", "", regex=False)
        .str.strip()
        .str.upper()
    )
    dfep = dfep.loc[:, ~dfep.columns.str.startswith("UNNAMED")]

    # ✅ clave única
    if "ID_EMPRESAS" not in dfep.columns:
        st.error("La hoja 'Empresas_Procedimientos' debe tener una columna 'ID_EMPRESAS' (clave única).")
        st.stop()

    dfep["ID_EMPRESAS"] = dfep["ID_EMPRESAS"].astype(str).str.strip()
    dfep = dfep[dfep["ID_EMPRESAS"] != ""]

    # =========================
    # FILTRO
    # =========================
    st.subheader("FILTRO")

    def _clean_series(s):
        return (
            s.dropna()
            .astype(str)
            .str.strip()
            .replace("", np.nan)
            .dropna()
        )

    opts_plan = ["TODOS"] + sorted(_clean_series(dfep["PLAN DE ACCIÓN ESPECÍFICO"]).unique().tolist())
    opts_acc = ["TODOS"] + sorted(_clean_series(dfep["ACCIONES FÍSICAS"]).unique().tolist())
    opts_proc = ["TODOS"] + sorted(_clean_series(dfep["PROCEDIMIENTO"]).unique().tolist())
    opts_esp = ["TODOS"] + sorted(_clean_series(dfep["ESPECIALISTA"]).unique().tolist())
    opts_emp = ["TODOS"] + sorted(_clean_series(dfep["EMPRESAS"]).unique().tolist())

    c1, c2, c3, c4, c5 = st.columns(5)
    with c1:
        f_plan = st.selectbox("PLAN DE ACCIÓN", opts_plan, key="f_ep_plan")
    with c2:
        f_acc = st.selectbox("ACCIONES FÍSICAS", opts_acc, key="f_ep_acc")
    with c3:
        f_proc = st.selectbox("PROCEDIMIENTO", opts_proc, key="f_ep_proc")
    with c4:
        f_esp = st.selectbox("ESPECIALISTA", opts_esp, key="f_ep_esp")
    with c5:
        f_emp = st.selectbox("EMPRESAS", opts_emp, key="f_ep_emp")

    # aplicar filtros
    dfep_f = dfep.copy()

    if f_plan != "TODOS":
        dfep_f = dfep_f[dfep_f["PLAN DE ACCIÓN ESPECÍFICO"].astype(str).str.strip() == f_plan]
    if f_acc != "TODOS":
        dfep_f = dfep_f[dfep_f["ACCIONES FÍSICAS"].astype(str).str.strip() == f_acc]
    if f_proc != "TODOS":
        dfep_f = dfep_f[dfep_f["PROCEDIMIENTO"].astype(str).str.strip() == f_proc]
    if f_esp != "TODOS":
        dfep_f = dfep_f[dfep_f["ESPECIALISTA"].astype(str).str.strip() == f_esp]
    if f_emp != "TODOS":
        dfep_f = dfep_f[dfep_f["EMPRESAS"].astype(str).str.strip() == f_emp]
    
    # =========================
    # RESUMEN (debajo del filtro)
    # =========================
    st.subheader("RESUMEN")
    
    # Empresas únicas (sin repetir)
    empresas_unicas = (
        dfep_f["EMPRESAS"]
        .dropna()
        .astype(str)
        .str.strip()
        .replace("", np.nan)
        .dropna()
        .nunique()
        if "EMPRESAS" in dfep_f.columns else 0
    )
    
    # Procedimientos únicos (sin repetir)
    procedimientos_unicos = (
        dfep_f["PROCEDIMIENTO"]
        .dropna()
        .astype(str)
        .str.strip()
        .replace("", np.nan)
        .dropna()
        .nunique()
        if "PROCEDIMIENTO" in dfep_f.columns else 0
    )
    
    r1, r2 = st.columns(2)
    with r1:
        st.metric("Empresas", int(empresas_unicas))
    with r2:
        st.metric("Procedimientos", int(procedimientos_unicos))
    
    # =========================
    # GRÁFICOS
    # =========================
    def get_plan_code(text):
        m = re.match(r"(\d+\.\d+)", str(text))
        return m.group(1) if m else text
    
    def _clean_col(df, col):
        return (
            df[col].dropna().astype(str).str.strip().replace("", np.nan).dropna()
            if col in df.columns else pd.Series([], dtype=str)
        )
    
    # 1) Procedimientos por Plan (únicos)
    proc_plan = (
        dfep_f[["PLAN DE ACCIÓN ESPECÍFICO", "PROCEDIMIENTO"]]
        .dropna()
        .assign(
            PLAN=lambda x: x["PLAN DE ACCIÓN ESPECÍFICO"].astype(str).str.strip(),
            PROC=lambda x: x["PROCEDIMIENTO"].astype(str).str.strip(),
        )
    )
    proc_plan = proc_plan[(proc_plan["PLAN"] != "") & (proc_plan["PROC"] != "")]
    proc_plan = (
        proc_plan.groupby("PLAN")["PROC"]
        .nunique()
        .reset_index(name="PROCEDIMIENTOS_UNICOS")
        .sort_values("PROCEDIMIENTOS_UNICOS", ascending=False)
    )
    
    # 2) Empresas por Plan (únicas)
    emp_plan = (
        dfep_f[["PLAN DE ACCIÓN ESPECÍFICO", "EMPRESAS"]]
        .dropna()
        .assign(
            PLAN=lambda x: x["PLAN DE ACCIÓN ESPECÍFICO"].astype(str).str.strip(),
            EMP=lambda x: x["EMPRESAS"].astype(str).str.strip(),
        )
    )
    emp_plan = emp_plan[(emp_plan["PLAN"] != "") & (emp_plan["EMP"] != "")]
    emp_plan = (
        emp_plan.groupby("PLAN")["EMP"]
        .nunique()
        .reset_index(name="EMPRESAS_UNICAS")
        .sort_values("EMPRESAS_UNICAS", ascending=False)
    )
    
    # 3) Especialistas por Plan (únicos)
    esp_plan = (
        dfep_f[["PLAN DE ACCIÓN ESPECÍFICO", "ESPECIALISTA"]]
        .dropna()
        .assign(
            PLAN=lambda x: x["PLAN DE ACCIÓN ESPECÍFICO"].astype(str).str.strip(),
            ESP=lambda x: x["ESPECIALISTA"].astype(str).str.strip(),
        )
    )
    esp_plan = esp_plan[(esp_plan["PLAN"] != "") & (esp_plan["ESP"] != "")]
    esp_plan = (
        esp_plan.groupby("PLAN")["ESP"]
        .nunique()
        .reset_index(name="ESPECIALISTAS_UNICOS")
        .sort_values("ESPECIALISTAS_UNICOS", ascending=False)
    )
    
    # 4) Procedimientos por Especialista (únicos)
    proc_esp = (
        dfep_f[["ESPECIALISTA", "PROCEDIMIENTO"]]
        .dropna()
        .assign(
            ESP=lambda x: x["ESPECIALISTA"].astype(str).str.strip(),
            PROC=lambda x: x["PROCEDIMIENTO"].astype(str).str.strip(),
        )
    )
    proc_esp = proc_esp[(proc_esp["ESP"] != "") & (proc_esp["PROC"] != "")]
    proc_esp = (
        proc_esp.groupby("ESP")["PROC"]
        .nunique()
        .reset_index(name="PROCEDIMIENTOS_UNICOS")
        .sort_values("PROCEDIMIENTOS_UNICOS", ascending=False)
    )
    
    # --- Helper para eje X en 2 líneas (2 filas) ---
    def wrap_x(name, max_chars=22):
        parts = str(name).split()
        if len(parts) <= 2:
            return name
        line1 = []
        for p in parts:
            if len(" ".join(line1 + [p])) <= max_chars:
                line1.append(p)
            else:
                break
        line2 = parts[len(line1):]
        if len(" ".join(line2)) < 8 and len(line1) > 1:
            line2.insert(0, line1.pop())
        return " ".join(line1) + "<br>" + " ".join(line2)
    
    # =========================
    # Layout 2x2
    # =========================
    g1, g2 = st.columns(2)
    g3, g4 = st.columns(2)
    
    # 1) Barras: Procedimientos por Plan
    with g1:
        tmp = proc_plan.copy()
        # Código corto (1.1, 1.2, ...)
        tmp["PLAN_CODE"] = tmp["PLAN"].apply(get_plan_code)
        # Texto largo SOLO para leyenda
        tmp["PLAN_FULL"] = tmp["PLAN"]
        # 🔑 ordenar por el código numérico
        tmp["_ORDEN"] = tmp["PLAN_CODE"].astype(str)
        tmp = tmp.sort_values("_ORDEN")
        
        orden_leyenda = tmp["PLAN_FULL"].tolist()
        
        fig1 = px.bar(
            tmp,
            x="PLAN_CODE",
            y="PROCEDIMIENTOS_UNICOS",
            color="PLAN_FULL",
            category_orders={
                "PLAN_FULL": orden_leyenda,   # 👈 fuerza orden de la leyenda
                "PLAN_CODE": tmp["PLAN_CODE"].tolist()
            },
            hover_data={
                "PLAN_CODE": False,
                "PLAN_FULL": True,
                "PROCEDIMIENTOS_UNICOS": True,
            },
            labels={
                "PLAN_CODE": "Plan de acción",
                "PROCEDIMIENTOS_UNICOS": "Procedimientos",
                "PLAN_FULL": "Plan de acción específico"
            },
            title="Procedimientos por Plan de Acción"
        )
        
        fig1.update_layout(
            xaxis_tickangle=0,
            legend=dict(
                orientation="v",
                y=-1.30,
                x=0.5,
                xanchor="center",
                font=dict(size=10),
                title_text=""
            )
        )
        
        st.plotly_chart(fig1, width="stretch")
        
    # 2) Treemap: Empresas por Plan
    with g2:
        fig2 = px.treemap(
            emp_plan,
            path=["PLAN"],
            values="EMPRESAS_UNICAS",
            hover_data=["EMPRESAS_UNICAS"],
            title="Empresas por Plan de Acción"
        )
        fig2.update_traces(
            textfont=dict(size=18),
            textinfo="label",
            hovertemplate="<b>%{label}</b><br>Empresas: %{value}<extra></extra>"
        )

        st.plotly_chart(fig2, width="stretch")
    
    # 3) Circular: Especialistas por Plan
    with g3:
        fig3 = px.pie(
            esp_plan,
            names="PLAN",
            values="ESPECIALISTAS_UNICOS",
            hover_data=["ESPECIALISTAS_UNICOS"],
            title="Especialistas por Plan de Acción"
        )
        
        fig3.update_layout(
            legend=dict(
                orientation="h",   # horizontal
                y=-0.25,           # debajo del gráfico
                x=0.5,
                xanchor="center",
                font=dict(size=10) # tamaño reducido
            )
        )
        
        st.plotly_chart(fig3, width="stretch")
    
    # 4) Barras: Procedimientos por Especialista
    with g4:
        tmp = proc_esp.copy()
        tmp["ESP_WRAP"] = tmp["ESP"].apply(lambda x: wrap_x(x, max_chars=18))
    
        fig4 = px.bar(
            tmp,
            x="ESP_WRAP",
            y="PROCEDIMIENTOS_UNICOS",
            hover_data=["PROCEDIMIENTOS_UNICOS"],
            labels={"ESP_WRAP": "Especialista", "PROCEDIMIENTOS_UNICOS": "Procedimientos"},
            title="Procedimientos por Especialista"
        )
        fig4.update_layout(xaxis_tickangle=0)
        st.plotly_chart(fig4, width="stretch")
    
    # =========================
    # TABLA EDITABLE
    # =========================  
    st.divider()
    st.subheader("TABLA EDITABLE")

    editable_cols_ep = [
        "PLAN DE ACCIÓN ESPECÍFICO",
        "ACCIONES FÍSICAS",
        "RCD/RD/PR/LEY",
        "PROCEDIMIENTO",
        "ESPECIALISTA",
        "EMPRESAS",
    ]

    missing = [c for c in editable_cols_ep if c not in dfep.columns]
    if missing:
        st.error(f"Faltan columnas en 'Empresas_Procedimientos': {missing}")
        st.stop()

    viewep = dfep_f[["ID_EMPRESAS"] + editable_cols_ep].copy()

    orig_key_ep = f"orig_empresas_procedimientos_{f_plan}_{f_acc}_{f_proc}_{f_esp}_{f_emp}"
    if orig_key_ep not in st.session_state:
        st.session_state[orig_key_ep] = viewep.copy()

    def _norm_ep(x):
        if x is None:
            return ""
        s = str(x)
        return "" if s.lower() == "nan" else s.strip()

    editedep = st.data_editor(
        viewep,
        width="stretch",
        hide_index=True,
        num_rows="fixed",
        key=f"editor_empresas_procedimientos_{f_plan}_{f_acc}_{f_proc}_{f_esp}_{f_emp}",
    )

    # =========================
    # GUARDAR CAMBIOS EN LOTE
    # =========================
    if st.button("💾 Guardar Empresa", key="save_empresa"):
        original = st.session_state[orig_key_ep].copy()

        changed = pd.Series(np.zeros(len(editedep), dtype=bool), index=editedep.index)
        for c in editable_cols_ep:
            changed |= editedep[c].map(_norm_ep) != original[c].map(_norm_ep)

        changed_rows = editedep.loc[changed].copy()

        if changed_rows.empty:
            st.info("No hay cambios.")
        else:
            updates = []
            for _, r in changed_rows.iterrows():
                fields = {c: _norm_ep(r[c]) for c in editable_cols_ep}
                updates.append({
                    "ID": _norm_ep(r["ID_EMPRESAS"]),
                    "FIELDS": fields
                })

            payload = {"action": "BATCH_UPDATE", "sheet": SHEET_NAME_EP, "updates": updates}
            resp = requests.post(APPS_SCRIPT_URL, json=payload, timeout=30)
            data = resp.json() if resp.ok else {"ok": False, "error": resp.text}

            if data.get("ok"):
                st.success(f"Guardado ✅ Actualizadas: {data.get('updated')} | No encontradas: {data.get('notFound')}")
                st.session_state[orig_key_ep] = editedep.copy()
                st.rerun()
            else:
                st.error(data)

    # =========================
    # AGREGAR FILA
    # =========================
    st.divider()
    st.subheader("AGREGAR FILA")

    with st.form("add_row_empresas_procedimientos"):
        plan = st.selectbox(
            "PLAN DE ACCIÓN ESPECÍFICO",
            [
                "1.1 Supervisión de la Generación del SEIN",
                "1.2 Supervisión del COES",
                "1.3 Supervisión de la Regulación del Mercado Eléctrico en los Agentes SEIN_COES",
                "1.4 Supervisión de la Operatividad de la Generación en Sistemas Eléctricos Aislados y la Generación Distribuida a cargo de las Empresas Distribuidoras y Autoproductoras",
                "1.5 Supervisión de las condiciones de la Generación Eléctrica para la prestación del servicio público de electricidad en los sistemas aislados y la Operatividad de los Sistemas Fotovoltaicos no conectados a Red",
                "1.6 Supervisión del Cumplimiento de la Norma Técnica de Calidad de los Servicios Eléctricos (NTCSE) y su base metodológicas",
            ],
            key="add_ep_plan",
        )

        acciones = st.selectbox(
            "ACCIONES FÍSICAS",
            [
                "Supervisión de la disponibilidad y operatividad unidades SEIN",
                "Supervisión ensayos unidades generación",
                "Supervisión del desempeño de las unidades de generación despachada por el COES",
                "Supervisión de las inflexibilidades operativas de las unidades de generación",
                "Supervisión programas de mantenimiento COES",
                "Supervisión de la programación de la operación",
                "Supervisión de coordinación de operación",
                "Supervisión evaluación de operación",
                "Supervisión esquemas rechazo carga generación",
                "Monitoreo operación en tiempo real",
                "Operación NETMECOTR y la NTIITR",
                "Operación semanal y margen de reserva",
                "Evaluación de procedimientos del COES",
                "Supervisión recaudación FISE - Generación",
                "Seguimiento de contratos de usuarios libres",
                "Supervisión costos incurridos de generación adicional",
                "Supervisión transacciones económicas",
                "Supervisión de generación aislada",
                "Supervisión de generación no despachada por el COES",
                "Supervisión de autoproductoras",
                "Supervisión de accidentes de terceros",
                "Supervisión de sistemas fotovoltaicos autónomos",
                "Supervisión de la Norma Técnica de calidad de los servicios eléctricos y su base metodológica",
            ],
            key="add_ep_acc",
        )

        rcd = st.text_input("RCD/RD/PR/LEY", key="add_ep_rcd")
        procedimiento = st.text_area("PROCEDIMIENTO", key="add_ep_proc")

        especialista = st.selectbox(
            "ESPECIALISTA",
            [
                "EDUARDO CARRILLO TINCALLPA",
                "JORGE ISRAEL MONTENEGRO SANTOS",
                "RUBEN ROJAS RAMIREZ",
                "JORGE PEDRO VILCACHAGUA NUÑEZ",
                "CESAR GUILLERMO OLANO OCHOA",
                "ANGEL DANIEL ROBLES SARAVIA",
                "GERMAN ABEL GUTARRA CRIBILLERO",
            ],
            key="add_ep_esp",
        )

        empresa = st.text_input("EMPRESA", key="add_ep_emp")
        add_ok_ep = st.form_submit_button("➕ Agregar")

    if add_ok_ep:
        payload = {
            "action": "ADD",
            "sheet": SHEET_NAME_EP,
            "row": {
                "PLAN DE ACCIÓN ESPECÍFICO": plan.strip(),
                "ACCIONES FÍSICAS": acciones.strip(),
                "RCD/RD/PR/LEY": rcd.strip(),
                "PROCEDIMIENTO": procedimiento.strip(),
                "ESPECIALISTA": especialista.strip(),
                "EMPRESAS": empresa.strip(),
            }
        }

        resp = requests.post(APPS_SCRIPT_URL, json=payload, timeout=30)
        data = resp.json() if resp.ok else {"ok": False, "error": resp.text}

        if data.get("ok"):
            st.success(f"Agregado ✅ ID_EMPRESAS = {data.get('id')}")
            st.rerun()
        else:
            st.error(data)

    # =========================
    # BORRAR FILA
    # =========================
    st.divider()
    st.subheader("BORRAR FILA")

    ids_borrado = dfep_f["ID_EMPRESAS"].astype(str).tolist() if not dfep_f.empty else dfep["ID_EMPRESAS"].astype(str).tolist()
    id_del = st.selectbox("ID_EMPRESAS a borrar", ids_borrado, key="del_ep")
    confirm_ep = st.checkbox("Confirmo borrado irreversible", key="conf_ep")

    if st.button("🗑️ Borrar", key="btn_del_ep"):
        if not confirm_ep:
            st.warning("Confirma antes de borrar.")
        else:
            payload = {"action": "DELETE", "sheet": SHEET_NAME_EP, "id": str(id_del)}
            resp = requests.post(APPS_SCRIPT_URL, json=payload, timeout=30)
            data = resp.json() if resp.ok else {"ok": False, "error": resp.text}

            if data.get("ok"):
                st.success("Borrado ✅")
                st.rerun()
            else:
                st.error(data)

# =========================
# TAB: SUP2500037
# =========================
with tab_SUP2500037:
    st.header("Contrato SUP2500037")

    SHEET_NAME_SUP = "SUP2500037"

    # =========================
    # LEER GOOGLE SHEET (CSV público)
    # =========================
    sheet_sup = urllib.parse.quote(SHEET_NAME_SUP)
    url_sup = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/gviz/tq?tqx=out:csv&sheet={sheet_sup}"

    dfsup = pd.read_csv(url_sup, dtype=str)
    dfsup.columns = (
        dfsup.columns.astype(str)
        .str.replace("\ufeff", "", regex=False)
        .str.strip()
        .str.upper()
    )
    dfsup = dfsup.loc[:, ~dfsup.columns.str.startswith("UNNAMED")]

    if "ID_SUP" not in dfsup.columns:
        st.error("La hoja 'SUP2500037' debe tener una columna 'ID_SUP' (clave única).")
        st.stop()

    dfsup["ID_SUP"] = dfsup["ID_SUP"].astype(str).str.strip()
    dfsup = dfsup[dfsup["ID_SUP"] != ""]

    # =========================
    # HELPERS
    # =========================
    def pick_col(df, candidates):
        cols = set(df.columns)
        for c in candidates:
            if c in cols:
                return c
        return None

    def _try_int(x):
        try:
            return int(str(x).strip())
        except:
            return None

    def clean_opts(s: pd.Series):
        s = s.dropna().astype(str).str.strip().replace("", np.nan).dropna()
        # ordenar numérico si aplica
        nums = pd.to_numeric(s, errors="coerce")
        if nums.notna().all() and len(nums) > 0:
            order = sorted(nums.unique().tolist())
            out = []
            for v in order:
                if float(v).is_integer():
                    out.append(str(int(v)))
                else:
                    out.append(str(v))
            return out
        return sorted(s.unique().tolist())

    def _eq(df, col, val):
        return df[col].astype(str).str.strip() == str(val).strip()

    # =========================
    # FILTRO (Año, Mes, Periodo, Especialista, Supervisor, Apoyo)
    # =========================
    st.subheader("FILTRO")

    col_anio = pick_col(dfsup, ["AÑO", "ANIO", "ANO", "YEAR"])
    col_mes = pick_col(dfsup, ["MES", "MONTH"])
    col_periodo = pick_col(dfsup, ["PERIODO", "PERÍODO", "PERIODO.", "PERÍODO."])
    col_esp = pick_col(dfsup, ["ESPECIALISTA"])
    col_sup = pick_col(dfsup, ["SUPERVISOR", "NOMBRE COMPLETO DE SUPERVISOR", "NOMBRE COMPLETO DEL SUPERVISOR"])
    col_apoyo = pick_col(dfsup, ["APOYO"])

    opts_anio = ["TODOS"] + (clean_opts(dfsup[col_anio]) if col_anio else [])
    opts_mes = ["TODOS"] + (clean_opts(dfsup[col_mes]) if col_mes else [])
    opts_periodo = ["TODOS"] + (clean_opts(dfsup[col_periodo]) if col_periodo else [])
    opts_esp = ["TODOS"] + (clean_opts(dfsup[col_esp]) if col_esp else [])
    opts_sup = ["TODOS"] + (clean_opts(dfsup[col_sup]) if col_sup else [])
    opts_apoyo = ["TODOS"] + (clean_opts(dfsup[col_apoyo]) if col_apoyo else [])

    c1, c2, c3 = st.columns(3)
    c4, c5, c6 = st.columns(3)

    with c1:
        f_anio = st.selectbox("Año", opts_anio, key="f_sup2500037_anio", disabled=(col_anio is None))
    with c2:
        f_mes = st.selectbox("Mes", opts_mes, key="f_sup2500037_mes", disabled=(col_mes is None))
    with c3:
        f_periodo = st.selectbox("Periodo", opts_periodo, key="f_sup2500037_periodo", disabled=(col_periodo is None))

    with c4:
        f_esp = st.selectbox("Especialista", opts_esp, key="f_sup2500037_esp", disabled=(col_esp is None))
    with c5:
        f_supervisor = st.selectbox("Supervisor", opts_sup, key="f_sup2500037_supervisor", disabled=(col_sup is None))
    with c6:
        f_apoyo = st.selectbox("Apoyo", opts_apoyo, key="f_sup2500037_apoyo", disabled=(col_apoyo is None))

    dfsup_f = dfsup.copy()
    if col_anio and f_anio != "TODOS":
        dfsup_f = dfsup_f[_eq(dfsup_f, col_anio, f_anio)]
    if col_mes and f_mes != "TODOS":
        dfsup_f = dfsup_f[_eq(dfsup_f, col_mes, f_mes)]
    if col_periodo and f_periodo != "TODOS":
        dfsup_f = dfsup_f[_eq(dfsup_f, col_periodo, f_periodo)]
    if col_esp and f_esp != "TODOS":
        dfsup_f = dfsup_f[_eq(dfsup_f, col_esp, f_esp)]
    if col_sup and f_supervisor != "TODOS":
        dfsup_f = dfsup_f[_eq(dfsup_f, col_sup, f_supervisor)]
    if col_apoyo and f_apoyo != "TODOS":
        dfsup_f = dfsup_f[_eq(dfsup_f, col_apoyo, f_apoyo)]
    
    # =========================
    # CARGA DESDE EXCEL (XLSX) -> AGREGA FILAS A SUP2500037 (CORREGIDO)
    # =========================
    st.divider()
    st.subheader("CARGAR EXCEL (.xlsx) Y AGREGAR FILAS")
    
    # -------------------------
    # INPUTS (meta)
    # -------------------------
    anio = st.selectbox("AÑO", list(range(2025, 2036)), index=0, key="up_sup_anio")
    
    mes = st.selectbox(
        "MES",
        ["ENERO","FEBRERO","MARZO","ABRIL","MAYO","JUNIO","JULIO","AGOSTO","SEPTIEMBRE","OCTUBRE","NOVIEMBRE","DICIEMBRE"],
        key="up_sup_mes",
    )
    
    periodo = st.text_input("PERIODO", key="up_sup_periodo")
    
    CONTRATO_AUTO = "SUP2500037"
    EST_AUTO = "VASMOL S.A.C."
    
    st.text_input("CONTRATO", value=CONTRATO_AUTO, disabled=False, key="up_sup_contrato")
    st.text_input("EST", value=EST_AUTO, disabled=False, key="up_sup_est")
    
    uploaded = st.file_uploader("Sube el Excel (.xlsx)", type=["xlsx"], key="up_sup_file")
    
    # -------------------------
    # NORMALIZACIÓN / HELPERS
    # -------------------------
    def strip_accents(s: str) -> str:
        return "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")
    
    def norm(s) -> str:
        if s is None:
            return ""
        s = str(s).replace("\n", " ").replace("\r", " ").strip()
        s = strip_accents(s).lower()
        s = re.sub(r"\s+", " ", s)
        return s
    
    def is_blank(v) -> bool:
        return v is None or (isinstance(v, str) and v.strip() == "")
    
    def row_is_empty(values) -> bool:
        return all(is_blank(v) for v in values)
    
    def safe_str(v) -> str:
        if v is None:
            return ""
        s = str(v).strip()
        return "" if s.lower() in ("none", "nan", "null") else s
    
    # -------------------------
    # CONFIG FILAS EXCEL
    # -------------------------
    HEADER_ROW_MAIN = 15
    HEADER_ROW_FALLBACK = 14
    DATA_START_ROW = 16
    
    # ✅ DESTINO = nombre EXACTO de columna en tu GOOGLE SHEET (SUP2500037)
    # ✅ ORIGEN  = cómo viene en el EXCEL subido (fila 14/15 combinadas o alternativas)
    # OJO: APOYO puede no existir -> se rellena vacío
    MAP_DEST_TO_SRC_ALTS = {
        "ALCANCE SEGÚN CONTRATO": ["Alcance según contrato"],
        "PROCEDIMIENTO / TEMA: DETALLE": ["Procedimiento / Tema", "Procedimiento / Tema: Detalle", "Detalle"],
        "PROCEDIMIENTO / TEMA: ACTIVIDAD A REALIZAR": ["Actividad a realizar", "Procedimiento / Tema: Actividad a realizar"],
        "UNIDAD OPERATIVA A SUPERVISAR: AGENTE SUPERVISADO": [
            "Unidad operativa a Supervisar / Fiscalizar: Empresa",
            "Unidad operativa a Supervisar: Agente Supervisado",
            "Agente Supervisado",
            "Empresa",
        ],
    
        # Opcionales:
        "UNIDAD OPERATIVA A SUPERVISAR: UNIDAD /EXPEDIENTE": ["Unidad operativa a Supervisar: Unidad /Expediente", "Unidad /Expediente", "Unidad / Expediente"],
        "UNIDAD OPERATIVA A SUPERVISAR: UBIGEO": ["Unidad operativa a Supervisar: Ubigeo", "Ubigeo"],
        "FECHA: EJECUCIÓN": ["Fecha: Ejecución", "Ejecución"],
        "FECHA: ENTREGABLE": ["Entrega de Informe", "Fecha: Entregable", "Entregable"],
        "ENTREGABLES": ["Observaciones", "Entregables", "Entregable"],
        "ESPECIALISTA": ["Especialista"],
        "SUPERVISOR": ["Supervisor"],
        "APOYO": ["Apoyo"], 
    }
        
    OPTIONAL_DESTS = {
        "UNIDAD OPERATIVA A SUPERVISAR: UNIDAD /EXPEDIENTE",
        "UNIDAD OPERATIVA A SUPERVISAR: UBIGEO",
        "FECHA: EJECUCIÓN",
        "ESPECIALISTA",
        "SUPERVISOR",
        "APOYO",
    }
    
    REQUIRED_DESTS = [k for k in MAP_DEST_TO_SRC_ALTS.keys() if k not in OPTIONAL_DESTS]
    
    # Columnas META en el Sheet (asegúrate que existan así en SUP2500037)
    META = {
        "AÑO": lambda: safe_str(anio),
        "MES": lambda: safe_str(mes),
        "PERIODO": lambda: safe_str(periodo),
        "CONTRATO": lambda: CONTRATO_AUTO,
        "EST": lambda: EST_AUTO,
    }
    
    def header_text(ws, col: int):
        """Combina fila14 y fila15 si ambas existen: 'Grupo: Subtitulo'."""
        v15 = ws.cell(HEADER_ROW_MAIN, col).value
        v14 = ws.cell(HEADER_ROW_FALLBACK, col).value
        v15s = "" if is_blank(v15) else str(v15).strip()
        v14s = "" if is_blank(v14) else str(v14).strip()
    
        if v14s and v15s:
            return f"{v14s}: {v15s}"
        return v15s or v14s
    
    def build_found(ws):
        """Mapa header_normalizado -> lista de cols."""
        found = {}
        for col in range(1, ws.max_column + 1):
            key = norm(header_text(ws, col))
            if key:
                found.setdefault(key, []).append(col)
        return found
    
    def score_sheet(found):
        score = 0
        for dest in REQUIRED_DESTS:
            ok = False
            for alt in MAP_DEST_TO_SRC_ALTS[dest]:
                if norm(alt) in found:
                    ok = True
                    break
            if ok:
                score += 1
        return score
    
    def resolve_columns(found):
        """Resuelve DESTINO -> col_excel (consumiendo columnas si hay duplicadas)."""
        found_local = {k: v.copy() for k, v in found.items()}
        dest_to_col = {}
        missing_required = []
    
        for dest, alts in MAP_DEST_TO_SRC_ALTS.items():
            chosen = None
            for alt in alts:
                k = norm(alt)
                if k in found_local and len(found_local[k]) > 0:
                    chosen = found_local[k].pop(0)
                    break
            
            if chosen is None and dest not in OPTIONAL_DESTS:
                missing_required.append((dest, alts))
                
            dest_to_col[dest] = chosen  # puede ser None (APOYO)
    
        return dest_to_col, missing_required
    
    def excel_to_rows(file_bytes: bytes):
        wb = load_workbook(BytesIO(file_bytes), data_only=True)
    
        # ✅ elegir automáticamente la hoja correcta (la de mayor score)
        best_ws = None
        best_found = None
        best_score = -1
    
        for ws in wb.worksheets:
            found = build_found(ws)
            sc = score_sheet(found)
            if sc > best_score:
                best_score = sc
                best_ws = ws
                best_found = found
    
        if best_ws is None or best_score < max(4, len(REQUIRED_DESTS) // 2):
            raise ValueError("No pude identificar la hoja correcta (encabezados no coinciden en ninguna hoja).")
    
        dest_to_col, missing_required = resolve_columns(best_found)
    
        if missing_required:
            msg = "Faltan encabezados requeridos en el Excel subido:\n"
            for dest, alts in missing_required:
                msg += f"- {dest} (busqué: {alts})\n"
            msg += "\nRevisa fila 14/15 o si el formato cambió."
            raise ValueError(msg)
    
        # ---- Diagnóstico: encabezados detectados ----
        diag = []
        for col in range(1, best_ws.max_column + 1):
            ht = header_text(best_ws, col)
            if ht:
                diag.append({"COL": col, "HEADER_DETECTADO": ht})
        diag_df = pd.DataFrame(diag)
    
        # ---- Extraer data desde fila 16 ----
        data = []
        r = DATA_START_ROW
        while True:
            row_vals = {}
            for dest, col in dest_to_col.items():
                if col is None:
                    row_vals[dest] = ""  # APOYO u opcional
                else:
                    row_vals[dest] = best_ws.cell(r, col).value
    
            if row_is_empty(list(row_vals.values())):
                break
    
            data.append(row_vals)
            r += 1
    
        if not data:
            raise ValueError("No se detectaron filas de datos (desde la fila 16).")
    
        # ---- REGLA A ----
        # Si están vacíos: DETALLE, EJECUCIÓN, ENTREGABLE, ENTREGABLES => copiar esas 4 de la fila anterior
        k_det = "PROCEDIMIENTO / TEMA: DETALLE"
        k_act = "PROCEDIMIENTO / TEMA: ACTIVIDAD A REALIZAR"
        k_eje = "FECHA: EJECUCIÓN"
        k_ent1 = "FECHA: ENTREGABLE"
        k_ent2 = "ENTREGABLES"
        k_alc = "ALCANCE SEGÚN CONTRATO"
    
        for i in range(1, len(data)):
            row = data[i]
            prev = data[i - 1]
            if is_blank(row[k_det]) and is_blank(row[k_eje]) and is_blank(row[k_ent1]) and is_blank(row[k_ent2]):
                row[k_det] = prev[k_det]
                row[k_eje] = prev[k_eje]
                row[k_ent1] = prev[k_ent1]
                row[k_ent2] = prev[k_ent2]
    
        # ---- REGLA B (forward-fill) ----
        ffill_keys = [k_alc, k_det, k_act]
        last_vals = {k: None for k in ffill_keys}
    
        for row in data:
            for k in ffill_keys:
                if is_blank(row[k]):
                    row[k] = last_vals[k]
                else:
                    last_vals[k] = row[k]
    
        # ---- Convertir a rows (keys EXACTAS del Sheet) ----
        rows = []
        for row in data:
            out = {}
    
            # columnas destino del sheet (exactas)
            for dest in MAP_DEST_TO_SRC_ALTS.keys():
                out[dest] = safe_str(row.get(dest, ""))
    
            # meta
            for k, fn in META.items():
                out[k] = safe_str(fn())
    
            rows.append(out)
    
        return rows, diag_df, best_ws.title, dest_to_col
    
    # =========================
    # UI: Procesar + Preview + Agregar
    # =========================
    if uploaded:
        try:
            rows_to_add, diag_df, sheet_used, dest_to_col = excel_to_rows(uploaded.getvalue())
    
            st.info(f"Hoja detectada en el Excel: **{sheet_used}**")
    
            with st.expander("🔍 Encabezados detectados (fila 14/15 combinadas)"):
                st.dataframe(diag_df, width="stretch", hide_index=True)
    
            with st.expander("🧭 Mapeo usado (DESTINO en Sheet → columna Excel)"):
                m = [{"DESTINO (SHEET)": k, "COL_EXCEL": v} for k, v in dest_to_col.items()]
                st.dataframe(pd.DataFrame(m), width="stretch", hide_index=True)
    
            st.markdown("**Vista previa (primeras 50 filas)**")
            prev_df = pd.DataFrame(rows_to_add).fillna("")
            st.dataframe(prev_df.head(50), width="stretch", hide_index=True)
            st.caption(f"Filas detectadas para agregar: {len(rows_to_add)}")
    
            confirm_add = st.checkbox("Confirmo que deseo agregar estas filas al final", key="confirm_add_xlsx_sup")
    
            if st.button("📥 Agregar filas al Sheet", key="btn_sup_add_xlsx"):
                if not confirm_add:
                    st.warning("Activa la confirmación antes de agregar.")
                    st.stop()
    
                if str(periodo).strip() == "":
                    st.error("PERIODO es obligatorio.")
                    st.stop()
    
                payload = {
                    "action": "BATCH_ADD",
                    "sheet": SHEET_NAME_SUP,   # "SUP2500037"
                    "rows": rows_to_add
                }
    
                resp = requests.post(APPS_SCRIPT_URL, json=payload, timeout=120)
                data = resp.json() if resp.ok else {"ok": False, "error": resp.text}
    
                if data.get("ok"):
                    st.success(
                        f"Agregado ✅ Filas: {data.get('added')} | "
                        f"ID_SUP: {data.get('id_first')}–{data.get('id_last')}"
                    )
                    st.rerun()
                else:
                    st.error(data)
    
        except Exception as ex:
            st.error(f"No se pudo procesar el Excel: {ex}")


            
    # =========================
    # TABLA EDITABLE
    # =========================
    st.divider()
    st.subheader("TABLA EDITABLE")

    editable_cols_sup = [c for c in dfsup.columns if c != "ID_SUP"]
    viewsup = dfsup_f[["ID_SUP"] + editable_cols_sup].copy()

    orig_key_sup = f"orig_sup2500037_{f_anio}_{f_mes}_{f_periodo}_{f_esp}_{f_supervisor}_{f_apoyo}"
    if orig_key_sup not in st.session_state:
        st.session_state[orig_key_sup] = viewsup.copy()

    def _norm_sup(x):
        if x is None:
            return ""
        s = str(x)
        return "" if s.lower() == "nan" else s.strip()

    editedsup = st.data_editor(
        viewsup,
        width="stretch",
        hide_index=True,
        num_rows="fixed",
        key=f"editor_sup2500037_{f_anio}_{f_mes}_{f_periodo}_{f_esp}_{f_supervisor}_{f_apoyo}",
    )

    if st.button("💾 Guardar SUP2500037", key="save_sup2500037"):
        original = st.session_state[orig_key_sup].copy()

        changed = pd.Series(False, index=editedsup.index)
        for c in editable_cols_sup:
            changed |= editedsup[c].map(_norm_sup) != original[c].map(_norm_sup)

        changed_rows = editedsup.loc[changed].copy()

        if changed_rows.empty:
            st.info("No hay cambios.")
        else:
            updates = []
            for _, r in changed_rows.iterrows():
                fields = {c: _norm_sup(r[c]) for c in editable_cols_sup}
                updates.append({"ID": _norm_sup(r["ID_SUP"]), "FIELDS": fields})

            payload = {"action": "BATCH_UPDATE", "sheet": SHEET_NAME_SUP, "updates": updates}
            resp = requests.post(APPS_SCRIPT_URL, json=payload, timeout=30)
            data = resp.json() if resp.ok else {"ok": False, "error": resp.text}

            if data.get("ok"):
                st.success(f"Guardado ✅ Filas actualizadas: {data.get('updated')} | No encontradas: {data.get('notFound')}")
                st.session_state[orig_key_sup] = editedsup.copy()
                st.rerun()
            else:
                st.error(data)

    # =========================
    # BORRAR FILA(S) + VISTA PREVIA (1 / Rango / Lista)
    # =========================
    st.divider()
    st.subheader("BORRAR FILA(S)")

    ids_sup = dfsup["ID_SUP"].dropna().astype(str).str.strip()
    ids_sup = ids_sup[ids_sup != ""].tolist()

    ids_int = [(_try_int(i), i) for i in ids_sup]
    if ids_int and all(v[0] is not None for v in ids_int):
        ids_sup_sorted = [v[1] for v in sorted(ids_int, key=lambda t: t[0])]
    else:
        ids_sup_sorted = sorted(ids_sup)

    modo = st.radio(
        "Modo de borrado",
        ["Uno", "Rango", "Lista"],
        horizontal=True,
        key="del_mode_sup2500037",
    )

    confirm = st.checkbox("Confirmo borrado irreversible", key="confirm_del_sup2500037")

    def preview_ids(target_ids):
        target_ids = [str(x).strip() for x in target_ids if str(x).strip() != ""]
        if not target_ids:
            return dfsup.head(0)  # vacío con todas columnas

        dfprev = dfsup.copy()
        dfprev["ID_SUP"] = dfprev["ID_SUP"].astype(str).str.strip()
        dfprev = dfprev[dfprev["ID_SUP"].isin(target_ids)]

        tmp = dfprev["ID_SUP"].map(_try_int)
        if tmp.notna().all() and len(tmp) > 0:
            dfprev = dfprev.assign(_ord=tmp).sort_values("_ord").drop(columns=["_ord"])

        cols = ["ID_SUP"] + [c for c in dfprev.columns if c != "ID_SUP"]
        return dfprev[cols]

    def delete_one(id_value: str):
        payload = {"action": "DELETE", "sheet": SHEET_NAME_SUP, "id": str(id_value)}
        resp = requests.post(APPS_SCRIPT_URL, json=payload, timeout=30)
        if not resp.ok:
            return False, resp.text
        data = resp.json()
        return bool(data.get("ok")), data

    if modo == "Uno":
        id_del = st.selectbox("ID_SUP a borrar", ids_sup_sorted, key="del_one_sup2500037")

        st.markdown("**Vista previa**")
        prev = preview_ids([id_del])
        if prev.empty:
            st.info("No se encontró la fila para ese ID_SUP.")
        else:
            st.dataframe(prev, width="stretch", hide_index=True)

        if st.button("🗑️ Borrar 1", key="btn_del_one_sup2500037"):
            if not confirm:
                st.warning("Confirma antes de borrar.")
            else:
                ok, data = delete_one(id_del)
                if ok:
                    st.success(f"Borrado ✅ ID_SUP={id_del}")
                    st.rerun()
                else:
                    st.error(data)

    elif modo == "Rango":
        colr1, colr2 = st.columns(2)
        with colr1:
            desde = st.text_input("Desde ID_SUP", key="del_range_from_sup2500037")
        with colr2:
            hasta = st.text_input("Hasta ID_SUP", key="del_range_to_sup2500037")

        st.caption("Ejemplo: desde 10 hasta 25 (incluye ambos).")

        a = _try_int(desde)
        b = _try_int(hasta)

        if desde.strip() != "" and hasta.strip() != "" and (a is None or b is None):
            st.error("Para rango, 'Desde' y 'Hasta' deben ser números (ID_SUP numérico).")

        if a is not None and b is not None:
            if a > b:
                a, b = b, a

            target_ids = [str(i) for i in range(a, b + 1)]

            st.markdown("**Vista previa**")
            prev = preview_ids(target_ids)
            if prev.empty:
                st.info("No se encontraron filas dentro del rango en la hoja.")
            else:
                st.dataframe(prev, width="stretch", hide_index=True)
                st.caption(f"Filas encontradas: {len(prev)} (IDs inexistentes se ignorarán).")

            if st.button("🗑️ Borrar rango", key="btn_del_range_sup2500037"):
                if not confirm:
                    st.warning("Confirma antes de borrar.")
                else:
                    borrados = 0
                    fallos = []
                    for tid in target_ids:
                        ok, data = delete_one(tid)
                        if ok:
                            borrados += 1
                        else:
                            fallos.append((tid, data))

                    if borrados:
                        st.success(f"Borrado ✅ {borrados} filas.")
                    if fallos:
                        st.warning(f"No se pudieron borrar {len(fallos)} IDs.")
                        st.write(fallos)

                    st.rerun()

    else:  # Lista
        ids_texto = st.text_area(
            "IDs a borrar (separados por coma o salto de línea)",
            placeholder="Ej: 1,3,7,20\n25\n30",
            key="del_list_sup2500037",
        )

        raw = ids_texto.replace("\n", ",")
        parts = [p.strip() for p in raw.split(",") if p.strip() != ""]

        seen = set()
        target_ids = []
        for p in parts:
            if p not in seen:
                seen.add(p)
                target_ids.append(p)

        st.markdown("**Vista previa**")
        prev = preview_ids(target_ids)
        if not target_ids:
            st.info("Ingresa IDs para ver vista previa.")
        elif prev.empty:
            st.info("No se encontraron filas con esos IDs en la hoja.")
        else:
            st.dataframe(prev, width="stretch", hide_index=True)
            st.caption(f"Filas encontradas: {len(prev)} (IDs inexistentes se ignorarán).")

        if st.button("🗑️ Borrar lista", key="btn_del_list_sup2500037"):
            if not confirm:
                st.warning("Confirma antes de borrar.")
            else:
                if not target_ids:
                    st.error("Ingresa al menos un ID.")
                    st.stop()

                borrados = 0
                fallos = []
                for tid in target_ids:
                    ok, data = delete_one(tid)
                    if ok:
                        borrados += 1
                    else:
                        fallos.append((tid, data))

                if borrados:
                    st.success(f"Borrado ✅ {borrados} filas.")
                if fallos:
                    st.warning(f"No se pudieron borrar {len(fallos)} IDs.")
                    st.write(fallos)

                st.rerun()

# =========================
# TAB: SUP2400128  
# =========================
with tab_SUP2400128:
    st.header("Contrato SUP2400128")

    SHEET_NAME_SUP = "SUP2400128"

    # =========================
    # LEER GOOGLE SHEET (CSV público)
    # =========================
    sheet_sup = urllib.parse.quote(SHEET_NAME_SUP)
    url_sup = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/gviz/tq?tqx=out:csv&sheet={sheet_sup}"

    dfsup = pd.read_csv(url_sup, dtype=str)
    dfsup.columns = (
        dfsup.columns.astype(str)
        .str.replace("\ufeff", "", regex=False)
        .str.strip()
        .str.upper()
    )
    dfsup = dfsup.loc[:, ~dfsup.columns.str.startswith("UNNAMED")]

    if "ID_SUP" not in dfsup.columns:
        st.error("La hoja 'SUP2400128' debe tener una columna 'ID_SUP' (clave única).")
        st.stop()

    dfsup["ID_SUP"] = dfsup["ID_SUP"].astype(str).str.strip()
    dfsup = dfsup[dfsup["ID_SUP"] != ""]
    dfsup = dfsup.fillna("")

    # =========================
    # HELPERS
    # =========================
    def pick_col(df, candidates):
        cols = set(df.columns)
        for c in candidates:
            if c in cols:
                return c
        return None

    def _try_int(x):
        try:
            return int(str(x).strip())
        except:
            return None

    def clean_opts(s: pd.Series):
        s = s.dropna().astype(str).str.strip().replace("", np.nan).dropna()
        nums = pd.to_numeric(s, errors="coerce")
        if nums.notna().all() and len(nums) > 0:
            order = sorted(nums.unique().tolist())
            out = []
            for v in order:
                if float(v).is_integer():
                    out.append(str(int(v)))
                else:
                    out.append(str(v))
            return out
        return sorted(s.unique().tolist())

    def _eq(df, col, val):
        return df[col].astype(str).str.strip() == str(val).strip()

    # =========================
    # FILTRO (Año, Mes, Periodo, Especialista, Supervisor, Apoyo)
    # =========================
    st.subheader("FILTRO")

    col_anio = pick_col(dfsup, ["AÑO", "ANIO", "ANO", "YEAR"])
    col_mes = pick_col(dfsup, ["MES", "MONTH"])
    col_periodo = pick_col(dfsup, ["PERIODO", "PERÍODO", "PERIODO.", "PERÍODO."])
    col_esp = pick_col(dfsup, ["ESPECIALISTA"])
    col_sup = pick_col(dfsup, ["SUPERVISOR", "NOMBRE COMPLETO DE SUPERVISOR", "NOMBRE COMPLETO DEL SUPERVISOR"])
    col_apoyo = pick_col(dfsup, ["APOYO"])

    opts_anio = ["TODOS"] + (clean_opts(dfsup[col_anio]) if col_anio else [])
    opts_mes = ["TODOS"] + (clean_opts(dfsup[col_mes]) if col_mes else [])
    opts_periodo = ["TODOS"] + (clean_opts(dfsup[col_periodo]) if col_periodo else [])
    opts_esp = ["TODOS"] + (clean_opts(dfsup[col_esp]) if col_esp else [])
    opts_sup = ["TODOS"] + (clean_opts(dfsup[col_sup]) if col_sup else [])
    opts_apoyo = ["TODOS"] + (clean_opts(dfsup[col_apoyo]) if col_apoyo else [])

    c1, c2, c3 = st.columns(3)
    c4, c5, c6 = st.columns(3)

    with c1:
        f_anio = st.selectbox("Año", opts_anio, key="f_sup2400128_anio", disabled=(col_anio is None))
    with c2:
        f_mes = st.selectbox("Mes", opts_mes, key="f_sup2400128_mes", disabled=(col_mes is None))
    with c3:
        f_periodo = st.selectbox("Periodo", opts_periodo, key="f_sup2400128_periodo", disabled=(col_periodo is None))

    with c4:
        f_esp = st.selectbox("Especialista", opts_esp, key="f_sup2400128_esp", disabled=(col_esp is None))
    with c5:
        f_supervisor = st.selectbox("Supervisor", opts_sup, key="f_sup2400128_supervisor", disabled=(col_sup is None))
    with c6:
        f_apoyo = st.selectbox("Apoyo", opts_apoyo, key="f_sup2400128_apoyo", disabled=(col_apoyo is None))

    dfsup_f = dfsup.copy()
    if col_anio and f_anio != "TODOS":
        dfsup_f = dfsup_f[_eq(dfsup_f, col_anio, f_anio)]
    if col_mes and f_mes != "TODOS":
        dfsup_f = dfsup_f[_eq(dfsup_f, col_mes, f_mes)]
    if col_periodo and f_periodo != "TODOS":
        dfsup_f = dfsup_f[_eq(dfsup_f, col_periodo, f_periodo)]
    if col_esp and f_esp != "TODOS":
        dfsup_f = dfsup_f[_eq(dfsup_f, col_esp, f_esp)]
    if col_sup and f_supervisor != "TODOS":
        dfsup_f = dfsup_f[_eq(dfsup_f, col_sup, f_supervisor)]
    if col_apoyo and f_apoyo != "TODOS":
        dfsup_f = dfsup_f[_eq(dfsup_f, col_apoyo, f_apoyo)]
    
    # =========================
    # CARGA DESDE EXCEL (XLSX) -> AGREGA FILAS A SUP2400128
    # =========================
    st.divider()
    st.subheader("CARGAR EXCEL (.xlsx) Y AGREGAR FILAS")

    # INPUTS (meta)
    anio_up = st.selectbox("AÑO", list(range(2025, 2036)), index=0, key="up_sup2400128_anio")
    mes_up = st.selectbox(
        "MES",
        ["ENERO","FEBRERO","MARZO","ABRIL","MAYO","JUNIO","JULIO","AGOSTO","SEPTIEMBRE","OCTUBRE","NOVIEMBRE","DICIEMBRE"],
        key="up_sup2400128_mes",
    )
    periodo_up = st.text_input("PERIODO", key="up_sup2400128_periodo")

    CONTRATO_AUTO = "SUP2400128"
    EST_AUTO = "Green Electricity Engineering Corporation Sucursal del Perú"

    st.text_input("CONTRATO", value=CONTRATO_AUTO, disabled=False, key="up_sup2400128_contrato")
    st.text_input("EST", value=EST_AUTO, disabled=False, key="up_sup2400128_est")

    uploaded = st.file_uploader("Sube el Excel (.xlsx)", type=["xlsx"], key="up_sup2400128_file")
    
    HEADER_ROW_MAIN = 13
    HEADER_ROW_FALLBACK = 12
    DATA_START_ROW = 14

    # DESTINO = headers del SHEET (SUP2300128) | ORIGEN = headers del Excel subido
    MAP_DEST_TO_SRC_ALTS = {
        "ALCANCE SEGÚN CONTRATO": ["Alcance según contrato"],
        "PROCEDIMIENTO / TEMA: DETALLE": ["Procedimiento / Tema", "Procedimiento / Tema: Detalle", "Detalle"],
        "PROCEDIMIENTO / TEMA: ACTIVIDAD A REALIZAR": ["Actividad a realizar", "Procedimiento / Tema: Actividad a realizar"],
        "UNIDAD OPERATIVA A SUPERVISAR: AGENTE SUPERVISADO": [
            "Unidad operativa a Supervisar / Fiscalizar: Empresa",
            "Unidad operativa a Supervisar: Agente Supervisado",
            "Agente Supervisado",
            "Empresa",
        ],
    
        # Opcionales:
        "UNIDAD OPERATIVA A SUPERVISAR: UNIDAD /EXPEDIENTE": ["Unidad operativa a Supervisar: Unidad /Expediente", "Unidad /Expediente", "Unidad / Expediente"],
        "UNIDAD OPERATIVA A SUPERVISAR: UBIGEO": ["Unidad operativa a Supervisar: Ubigeo", "Ubigeo"],
        "FECHA: EJECUCIÓN": ["Fecha: Ejecución", "Ejecución"],
        "FECHA: ENTREGABLE": ["Entrega de Informe", "Fecha: Entregable", "Entregable"],
        "ENTREGABLES": ["Observaciones", "Entregables", "Entregable"],
        "ESPECIALISTA": ["Especialista"],
        "SUPERVISOR": ["Supervisor"],
        "APOYO": ["Apoyo"], 
    }
    
    OPTIONAL_DESTS = {
        "UNIDAD OPERATIVA A SUPERVISAR: UNIDAD /EXPEDIENTE",
        "UNIDAD OPERATIVA A SUPERVISAR: UBIGEO",
        "FECHA: EJECUCIÓN",
        "ESPECIALISTA",
        "SUPERVISOR",
        "APOYO",
    }
    
    REQUIRED_DESTS = [k for k in MAP_DEST_TO_SRC_ALTS.keys() if k not in OPTIONAL_DESTS] 

    META = {
        "AÑO": lambda: safe_str(anio_up),
        "MES": lambda: safe_str(mes_up),
        "PERIODO": lambda: safe_str(periodo_up),
        "CONTRATO": lambda: CONTRATO_AUTO,
        "EST": lambda: EST_AUTO,
    }

    def excel_to_rows(file_bytes: bytes):
        wb = load_workbook(BytesIO(file_bytes), data_only=True)

        best_ws = None
        best_found = None
        best_score = -1
        for ws in wb.worksheets:
            found = build_found(ws)
            sc = score_sheet(found)
            if sc > best_score:
                best_score = sc
                best_ws = ws
                best_found = found

        if best_ws is None or best_score < max(4, len(REQUIRED_DESTS) // 2):
            raise ValueError("No pude identificar la hoja correcta (encabezados no coinciden en ninguna hoja).")

        dest_to_col, missing_required = resolve_columns(best_found)
        if missing_required:
            msg = "Faltan encabezados requeridos en el Excel subido:\n"
            for dest, alts in missing_required:
                msg += f"- {dest} (busqué: {alts})\n"
            raise ValueError(msg)

        diag = []
        for col in range(1, best_ws.max_column + 1):
            ht = header_text(best_ws, col)
            if ht:
                diag.append({"COL": col, "HEADER_DETECTADO": ht})
        diag_df = pd.DataFrame(diag)

        data = []
        r = DATA_START_ROW
        while True:
            row_vals = {}
            for dest, col in dest_to_col.items():
                row_vals[dest] = "" if col is None else best_ws.cell(r, col).value
            if row_is_empty(list(row_vals.values())):
                break
            data.append(row_vals)
            r += 1

        if not data:
            raise ValueError("No se detectaron filas de datos (desde la fila 14).")

        # Regla A
        k_det = "PROCEDIMIENTO / TEMA: DETALLE"
        k_act = "PROCEDIMIENTO / TEMA: ACTIVIDAD A REALIZAR"
        k_eje = "FECHA: EJECUCIÓN"
        k_ent1 = "FECHA: ENTREGABLE"
        k_ent2 = "ENTREGABLES"
        k_alc = "ALCANCE SEGÚN CONTRATO"

        for i in range(1, len(data)):
            row = data[i]
            prev = data[i - 1]
            if is_blank(row[k_det]) and is_blank(row[k_eje]) and is_blank(row[k_ent1]) and is_blank(row[k_ent2]):
                row[k_det] = prev[k_det]
                row[k_eje] = prev[k_eje]
                row[k_ent1] = prev[k_ent1]
                row[k_ent2] = prev[k_ent2]

        # Regla B
        ffill_keys = [k_alc, k_det, k_act]
        last_vals = {k: None for k in ffill_keys}
        for row in data:
            for k in ffill_keys:
                if is_blank(row[k]):
                    row[k] = last_vals[k]
                else:
                    last_vals[k] = row[k]

        rows = []
        for row in data:
            out = {}
            for dest in MAP_DEST_TO_SRC_ALTS.keys():
                out[dest] = safe_str(row.get(dest, ""))
            for k, fn in META.items():
                out[k] = safe_str(fn())
            rows.append(out)

        return rows, diag_df, best_ws.title, dest_to_col

    if uploaded:
        try:
            rows_to_add, diag_df, sheet_used, dest_to_col = excel_to_rows(uploaded.getvalue())

            st.info(f"Hoja detectada en el Excel: **{sheet_used}**")

            with st.expander("🔍 Encabezados detectados (fila 12/13 combinadas)"):
                st.dataframe(diag_df, width="stretch", hide_index=True)

            with st.expander("🧭 Mapeo usado (DESTINO en Sheet → columna Excel)"):
                m = [{"DESTINO (SHEET)": k, "COL_EXCEL": v} for k, v in dest_to_col.items()]
                st.dataframe(pd.DataFrame(m), width="stretch", hide_index=True)

            st.markdown("**Vista previa (primeras 50 filas)**")
            prev_df = pd.DataFrame(rows_to_add).fillna("")
            st.dataframe(prev_df.head(50), width="stretch", hide_index=True)
            st.caption(f"Filas detectadas para agregar: {len(rows_to_add)}")

            confirm_add = st.checkbox("Confirmo que deseo agregar estas filas al final", key="confirm_add_xlsx_sup2400128")

            if st.button("📥 Agregar filas al Sheet", key="btn_sup2400128_add_xlsx"):
                if not confirm_add:
                    st.warning("Activa la confirmación antes de agregar.")
                    st.stop()

                if str(periodo_up).strip() == "":
                    st.error("PERIODO es obligatorio.")
                    st.stop()

                payload = {
                    "action": "BATCH_ADD",
                    "sheet": SHEET_NAME_SUP,   # "SUP2400128"
                    "rows": rows_to_add
                }

                resp = requests.post(APPS_SCRIPT_URL, json=payload, timeout=120)
                data = resp.json() if resp.ok else {"ok": False, "error": resp.text}

                if data.get("ok"):
                    st.success(
                        f"Agregado ✅ Filas: {data.get('added')} | "
                        f"ID_SUP: {data.get('id_first')}–{data.get('id_last')}"
                    )
                    st.rerun()
                else:
                    st.error(data)

        except Exception as ex:
            st.error(f"No se pudo procesar el Excel: {ex}")

    # =========================
    # TABLA EDITABLE
    # =========================
    st.divider()
    st.subheader("TABLA EDITABLE")

    editable_cols_sup = [c for c in dfsup.columns if c != "ID_SUP"]
    viewsup = dfsup_f[["ID_SUP"] + editable_cols_sup].copy()

    orig_key_sup = f"orig_sup2400128_{f_anio}_{f_mes}_{f_periodo}_{f_esp}_{f_supervisor}_{f_apoyo}"
    if orig_key_sup not in st.session_state:
        st.session_state[orig_key_sup] = viewsup.copy()

    def _norm_sup(x):
        if x is None:
            return ""
        s = str(x)
        return "" if s.lower() in ("nan", "none", "null") else s.strip()

    editedsup = st.data_editor(
        viewsup,
        width="stretch",
        hide_index=True,
        num_rows="fixed",
        key=f"editor_sup2400128_{f_anio}_{f_mes}_{f_periodo}_{f_esp}_{f_supervisor}_{f_apoyo}",
    )

    if st.button("💾 Guardar SUP2400128", key="save_sup2400128"):
        original = st.session_state[orig_key_sup].copy()

        changed = pd.Series(False, index=editedsup.index)
        for c in editable_cols_sup:
            changed |= editedsup[c].map(_norm_sup) != original[c].map(_norm_sup)

        changed_rows = editedsup.loc[changed].copy()

        if changed_rows.empty:
            st.info("No hay cambios.")
        else:
            updates = []
            for _, r in changed_rows.iterrows():
                fields = {c: _norm_sup(r[c]) for c in editable_cols_sup}
                updates.append({"ID": _norm_sup(r["ID_SUP"]), "FIELDS": fields})

            payload = {"action": "BATCH_UPDATE", "sheet": SHEET_NAME_SUP, "updates": updates}
            resp = requests.post(APPS_SCRIPT_URL, json=payload, timeout=30)
            data = resp.json() if resp.ok else {"ok": False, "error": resp.text}

            if data.get("ok"):
                st.success(f"Guardado ✅ Filas actualizadas: {data.get('updated')} | No encontradas: {data.get('notFound')}")
                st.session_state[orig_key_sup] = editedsup.copy()
                st.rerun()
            else:
                st.error(data)

    # =========================
    # BORRAR FILA(S) + VISTA PREVIA (1 / Rango / Lista)
    # =========================
    st.divider()
    st.subheader("BORRAR FILA(S)")

    ids_sup = dfsup["ID_SUP"].dropna().astype(str).str.strip()
    ids_sup = ids_sup[ids_sup != ""].tolist()

    ids_int = [(_try_int(i), i) for i in ids_sup]
    if ids_int and all(v[0] is not None for v in ids_int):
        ids_sup_sorted = [v[1] for v in sorted(ids_int, key=lambda t: t[0])]
    else:
        ids_sup_sorted = sorted(ids_sup)

    modo = st.radio(
        "Modo de borrado",
        ["Uno", "Rango", "Lista"],
        horizontal=True,
        key="del_mode_sup2400128",
    )

    confirm = st.checkbox("Confirmo borrado irreversible", key="confirm_del_sup2400128")

    def preview_ids(target_ids):
        target_ids = [str(x).strip() for x in target_ids if str(x).strip() != ""]
        if not target_ids:
            return dfsup.head(0)

        dfprev = dfsup.copy()
        dfprev["ID_SUP"] = dfprev["ID_SUP"].astype(str).str.strip()
        dfprev = dfprev[dfprev["ID_SUP"].isin(target_ids)]

        tmp = dfprev["ID_SUP"].map(_try_int)
        if tmp.notna().all() and len(tmp) > 0:
            dfprev = dfprev.assign(_ord=tmp).sort_values("_ord").drop(columns=["_ord"])

        cols = ["ID_SUP"] + [c for c in dfprev.columns if c != "ID_SUP"]
        return dfprev[cols]

    def delete_one(id_value: str):
        payload = {"action": "DELETE", "sheet": SHEET_NAME_SUP, "id": str(id_value)}
        resp = requests.post(APPS_SCRIPT_URL, json=payload, timeout=30)
        if not resp.ok:
            return False, resp.text
        data = resp.json()
        return bool(data.get("ok")), data

    if modo == "Uno":
        id_del = st.selectbox("ID_SUP a borrar", ids_sup_sorted, key="del_one_sup2400128")

        st.markdown("**Vista previa**")
        prev = preview_ids([id_del])
        if prev.empty:
            st.info("No se encontró la fila para ese ID_SUP.")
        else:
            st.dataframe(prev, width="stretch", hide_index=True)

        if st.button("🗑️ Borrar 1", key="btn_del_one_sup2400128"):
            if not confirm:
                st.warning("Confirma antes de borrar.")
            else:
                ok, data = delete_one(id_del)
                if ok:
                    st.success(f"Borrado ✅ ID_SUP={id_del}")
                    st.rerun()
                else:
                    st.error(data)

    elif modo == "Rango":
        colr1, colr2 = st.columns(2)
        with colr1:
            desde = st.text_input("Desde ID_SUP", key="del_range_from_sup2400128")
        with colr2:
            hasta = st.text_input("Hasta ID_SUP", key="del_range_to_sup2400128")

        st.caption("Ejemplo: desde 10 hasta 25 (incluye ambos).")

        a = _try_int(desde)
        b = _try_int(hasta)

        if desde.strip() != "" and hasta.strip() != "" and (a is None or b is None):
            st.error("Para rango, 'Desde' y 'Hasta' deben ser números (ID_SUP numérico).")

        if a is not None and b is not None:
            if a > b:
                a, b = b, a

            target_ids = [str(i) for i in range(a, b + 1)]

            st.markdown("**Vista previa**")
            prev = preview_ids(target_ids)
            if prev.empty:
                st.info("No se encontraron filas dentro del rango en la hoja.")
            else:
                st.dataframe(prev, width="stretch", hide_index=True)
                st.caption(f"Filas encontradas: {len(prev)} (IDs inexistentes se ignorarán).")

            if st.button("🗑️ Borrar rango", key="btn_del_range_sup2400128"):
                if not confirm:
                    st.warning("Confirma antes de borrar.")
                else:
                    borrados = 0
                    fallos = []
                    for tid in target_ids:
                        ok, data = delete_one(tid)
                        if ok:
                            borrados += 1
                        else:
                            fallos.append((tid, data))

                    if borrados:
                        st.success(f"Borrado ✅ {borrados} filas.")
                    if fallos:
                        st.warning(f"No se pudieron borrar {len(fallos)} IDs.")
                        st.write(fallos)

                    st.rerun()

    else:  # Lista
        ids_texto = st.text_area(
            "IDs a borrar (separados por coma o salto de línea)",
            placeholder="Ej: 1,3,7,20\n25\n30",
            key="del_list_sup2400128",
        )

        raw = ids_texto.replace("\n", ",")
        parts = [p.strip() for p in raw.split(",") if p.strip() != ""]

        seen = set()
        target_ids = []
        for p in parts:
            if p not in seen:
                seen.add(p)
                target_ids.append(p)

        st.markdown("**Vista previa**")
        prev = preview_ids(target_ids)
        if not target_ids:
            st.info("Ingresa IDs para ver vista previa.")
        elif prev.empty:
            st.info("No se encontraron filas con esos IDs en la hoja.")
        else:
            st.dataframe(prev, width="stretch", hide_index=True)
            st.caption(f"Filas encontradas: {len(prev)} (IDs inexistentes se ignorarán).")

        if st.button("🗑️ Borrar lista", key="btn_del_list_sup2400128"):
            if not confirm:
                st.warning("Confirma antes de borrar.")
            else:
                if not target_ids:
                    st.error("Ingresa al menos un ID.")
                    st.stop()

                borrados = 0
                fallos = []
                for tid in target_ids:
                    ok, data = delete_one(tid)
                    if ok:
                        borrados += 1
                    else:
                        fallos.append((tid, data))

                if borrados:
                    st.success(f"Borrado ✅ {borrados} filas.")
                if fallos:
                    st.warning(f"No se pudieron borrar {len(fallos)} IDs.")
                    st.write(fallos)

                st.rerun()

# =========================
# TAB: SUP2400205 
# =========================
with tab_SUP2400205:
    st.header("Contrato SUP2400205")

    SHEET_NAME_SUP = "SUP2400205"

    # =========================
    # LEER GOOGLE SHEET (CSV público)
    # =========================
    sheet_sup = urllib.parse.quote(SHEET_NAME_SUP)
    url_sup = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/gviz/tq?tqx=out:csv&sheet={sheet_sup}"

    dfsup = pd.read_csv(url_sup, dtype=str)
    dfsup.columns = (
        dfsup.columns.astype(str)
        .str.replace("\ufeff", "", regex=False)
        .str.strip()
        .str.upper()
    )
    dfsup = dfsup.loc[:, ~dfsup.columns.str.startswith("UNNAMED")]

    if "ID_SUP" not in dfsup.columns:
        st.error("La hoja 'SUP2400205' debe tener una columna 'ID_SUP' (clave única).")
        st.stop()

    dfsup["ID_SUP"] = dfsup["ID_SUP"].astype(str).str.strip()
    dfsup = dfsup[dfsup["ID_SUP"] != ""]
    dfsup = dfsup.fillna("")

    # =========================
    # HELPERS
    # =========================
    def pick_col(df, candidates):
        cols = set(df.columns)
        for c in candidates:
            if c in cols:
                return c
        return None

    def _try_int(x):
        try:
            return int(str(x).strip())
        except:
            return None

    def clean_opts(s: pd.Series):
        s = s.dropna().astype(str).str.strip().replace("", np.nan).dropna()
        nums = pd.to_numeric(s, errors="coerce")
        if nums.notna().all() and len(nums) > 0:
            order = sorted(nums.unique().tolist())
            out = []
            for v in order:
                if float(v).is_integer():
                    out.append(str(int(v)))
                else:
                    out.append(str(v))
            return out
        return sorted(s.unique().tolist())

    def _eq(df, col, val):
        return df[col].astype(str).str.strip() == str(val).strip()

    # =========================
    # FILTRO (Año, Mes, Periodo, Especialista, Supervisor, Apoyo)
    # =========================
    st.subheader("FILTRO")

    col_anio = pick_col(dfsup, ["AÑO", "ANIO", "ANO", "YEAR"])
    col_mes = pick_col(dfsup, ["MES", "MONTH"])
    col_periodo = pick_col(dfsup, ["PERIODO", "PERÍODO", "PERIODO.", "PERÍODO."])
    col_esp = pick_col(dfsup, ["ESPECIALISTA"])
    col_sup = pick_col(dfsup, ["SUPERVISOR", "NOMBRE COMPLETO DE SUPERVISOR", "NOMBRE COMPLETO DEL SUPERVISOR"])
    col_apoyo = pick_col(dfsup, ["APOYO"])

    opts_anio = ["TODOS"] + (clean_opts(dfsup[col_anio]) if col_anio else [])
    opts_mes = ["TODOS"] + (clean_opts(dfsup[col_mes]) if col_mes else [])
    opts_periodo = ["TODOS"] + (clean_opts(dfsup[col_periodo]) if col_periodo else [])
    opts_esp = ["TODOS"] + (clean_opts(dfsup[col_esp]) if col_esp else [])
    opts_sup = ["TODOS"] + (clean_opts(dfsup[col_sup]) if col_sup else [])
    opts_apoyo = ["TODOS"] + (clean_opts(dfsup[col_apoyo]) if col_apoyo else [])

    c1, c2, c3 = st.columns(3)
    c4, c5, c6 = st.columns(3)

    with c1:
        f_anio = st.selectbox("Año", opts_anio, key="f_sup2400205_anio", disabled=(col_anio is None))
    with c2:
        f_mes = st.selectbox("Mes", opts_mes, key="f_sup2400205_mes", disabled=(col_mes is None))
    with c3:
        f_periodo = st.selectbox("Periodo", opts_periodo, key="f_sup2400205_periodo", disabled=(col_periodo is None))

    with c4:
        f_esp = st.selectbox("Especialista", opts_esp, key="f_sup2400205_esp", disabled=(col_esp is None))
    with c5:
        f_supervisor = st.selectbox("Supervisor", opts_sup, key="f_sup2400205_supervisor", disabled=(col_sup is None))
    with c6:
        f_apoyo = st.selectbox("Apoyo", opts_apoyo, key="f_sup2400205_apoyo", disabled=(col_apoyo is None))

    dfsup_f = dfsup.copy()
    if col_anio and f_anio != "TODOS":
        dfsup_f = dfsup_f[_eq(dfsup_f, col_anio, f_anio)]
    if col_mes and f_mes != "TODOS":
        dfsup_f = dfsup_f[_eq(dfsup_f, col_mes, f_mes)]
    if col_periodo and f_periodo != "TODOS":
        dfsup_f = dfsup_f[_eq(dfsup_f, col_periodo, f_periodo)]
    if col_esp and f_esp != "TODOS":
        dfsup_f = dfsup_f[_eq(dfsup_f, col_esp, f_esp)]
    if col_sup and f_supervisor != "TODOS":
        dfsup_f = dfsup_f[_eq(dfsup_f, col_sup, f_supervisor)]
    if col_apoyo and f_apoyo != "TODOS":
        dfsup_f = dfsup_f[_eq(dfsup_f, col_apoyo, f_apoyo)]
    
    # =========================
    # CARGA DESDE EXCEL (XLSX) -> AGREGA FILAS A SUP2300128
    # =========================
    st.divider()
    st.subheader("CARGAR EXCEL (.xlsx) Y AGREGAR FILAS")

    # INPUTS (meta)
    anio_up = st.selectbox("AÑO", list(range(2025, 2036)), index=0, key="up_sup2400205_anio")
    mes_up = st.selectbox(
        "MES",
        ["ENERO","FEBRERO","MARZO","ABRIL","MAYO","JUNIO","JULIO","AGOSTO","SEPTIEMBRE","OCTUBRE","NOVIEMBRE","DICIEMBRE"],
        key="up_sup2400205_mes",
    )
    periodo_up = st.text_input("PERIODO", key="up_sup2400205_periodo")

    CONTRATO_AUTO = "SUP2400205"
    EST_AUTO = "Vision Quality Energy S.A.C."

    st.text_input("CONTRATO", value=CONTRATO_AUTO, disabled=False, key="up_sup2400205_contrato")
    st.text_input("EST", value=EST_AUTO, disabled=False, key="up_sup2400205_est")

    uploaded = st.file_uploader("Sube el Excel (.xlsx)", type=["xlsx"], key="up_sup2400205_file")
    
    HEADER_ROW_MAIN = 13
    HEADER_ROW_FALLBACK = 12
    DATA_START_ROW = 14

    # DESTINO = headers del SHEET (SUP2400205) | ORIGEN = headers del Excel subido
    MAP_DEST_TO_SRC_ALTS = {
        "ALCANCE SEGÚN CONTRATO": ["Alcance según contrato"],
        "PROCEDIMIENTO / TEMA: DETALLE": ["Procedimiento / Tema", "Procedimiento / Tema: Detalle", "Detalle"],
        "PROCEDIMIENTO / TEMA: ACTIVIDAD A REALIZAR": ["Actividad a realizar", "Procedimiento / Tema: Actividad a realizar"],
        "UNIDAD OPERATIVA A SUPERVISAR: AGENTE SUPERVISADO": [
            "Unidad operativa a Supervisar / Fiscalizar: Empresa",
            "Unidad operativa a Supervisar: Agente Supervisado",
            "Agente Supervisado",
            "Empresa",
        ],
    
        # Opcionales:
        "UNIDAD OPERATIVA A SUPERVISAR: UNIDAD /EXPEDIENTE": ["Unidad operativa a Supervisar: Unidad /Expediente", "Unidad /Expediente", "Unidad / Expediente"],
        "UNIDAD OPERATIVA A SUPERVISAR: UBIGEO": ["Unidad operativa a Supervisar: Ubigeo", "Ubigeo"],
        "FECHA: EJECUCIÓN": ["Fecha: Ejecución", "Ejecución"],
        "FECHA: ENTREGABLE": ["Entrega de Informe", "Fecha: Entregable", "Entregable"],
        "ENTREGABLES": ["Observaciones", "Entregables", "Entregable"],
        "ESPECIALISTA": ["Especialista"],
        "SUPERVISOR": ["Supervisor"],
        "APOYO": ["Apoyo"], 
    }

    
    OPTIONAL_DESTS = {
        "UNIDAD OPERATIVA A SUPERVISAR: UNIDAD /EXPEDIENTE",
        "UNIDAD OPERATIVA A SUPERVISAR: UBIGEO",
        "FECHA: EJECUCIÓN",
        "ESPECIALISTA",
        "SUPERVISOR",
        "APOYO",
    }
    
    REQUIRED_DESTS = [k for k in MAP_DEST_TO_SRC_ALTS.keys() if k not in OPTIONAL_DESTS]
    
    META = {
        "AÑO": lambda: safe_str(anio_up),
        "MES": lambda: safe_str(mes_up),
        "PERIODO": lambda: safe_str(periodo_up),
        "CONTRATO": lambda: CONTRATO_AUTO,
        "EST": lambda: EST_AUTO,
    }
    
    def excel_to_rows(file_bytes: bytes):
        wb = load_workbook(BytesIO(file_bytes), data_only=True)

        best_ws = None
        best_found = None
        best_score = -1
        for ws in wb.worksheets:
            found = build_found(ws)
            sc = score_sheet(found)
            if sc > best_score:
                best_score = sc
                best_ws = ws
                best_found = found

        if best_ws is None or best_score < max(4, len(REQUIRED_DESTS) // 2):
            raise ValueError("No pude identificar la hoja correcta (encabezados no coinciden en ninguna hoja).")

        dest_to_col, missing_required = resolve_columns(best_found)
        if missing_required:
            msg = "Faltan encabezados requeridos en el Excel subido:\n"
            for dest, alts in missing_required:
                msg += f"- {dest} (busqué: {alts})\n"
            raise ValueError(msg)

        diag = []
        for col in range(1, best_ws.max_column + 1):
            ht = header_text(best_ws, col)
            if ht:
                diag.append({"COL": col, "HEADER_DETECTADO": ht})
        diag_df = pd.DataFrame(diag)

        data = []
        r = DATA_START_ROW
        while True:
            row_vals = {}
            for dest, col in dest_to_col.items():
                row_vals[dest] = "" if col is None else best_ws.cell(r, col).value
            if row_is_empty(list(row_vals.values())):
                break
            data.append(row_vals)
            r += 1

        if not data:
            raise ValueError("No se detectaron filas de datos (desde la fila 14).")

        # Regla A
        k_det = "PROCEDIMIENTO / TEMA: DETALLE"
        k_act = "PROCEDIMIENTO / TEMA: ACTIVIDAD A REALIZAR"
        k_eje = "FECHA: EJECUCIÓN"
        k_ent1 = "FECHA: ENTREGABLE"
        k_ent2 = "ENTREGABLES"
        k_alc = "ALCANCE SEGÚN CONTRATO"

        for i in range(1, len(data)):
            row = data[i]
            prev = data[i - 1]
            if is_blank(row[k_det]) and is_blank(row[k_eje]) and is_blank(row[k_ent1]) and is_blank(row[k_ent2]):
                row[k_det] = prev[k_det]
                row[k_eje] = prev[k_eje]
                row[k_ent1] = prev[k_ent1]
                row[k_ent2] = prev[k_ent2]

        # Regla B
        ffill_keys = [k_alc, k_det, k_act]
        last_vals = {k: None for k in ffill_keys}
        for row in data:
            for k in ffill_keys:
                if is_blank(row[k]):
                    row[k] = last_vals[k]
                else:
                    last_vals[k] = row[k]

        rows = []
        for row in data:
            out = {}
            for dest in MAP_DEST_TO_SRC_ALTS.keys():
                out[dest] = safe_str(row.get(dest, ""))
            for k, fn in META.items():
                out[k] = safe_str(fn())
            rows.append(out)

        return rows, diag_df, best_ws.title, dest_to_col

    if uploaded:
        try:
            rows_to_add, diag_df, sheet_used, dest_to_col = excel_to_rows(uploaded.getvalue())

            st.info(f"Hoja detectada en el Excel: **{sheet_used}**")

            with st.expander("🔍 Encabezados detectados (fila 12/13 combinadas)"):
                st.dataframe(diag_df, width="stretch", hide_index=True)

            with st.expander("🧭 Mapeo usado (DESTINO en Sheet → columna Excel)"):
                m = [{"DESTINO (SHEET)": k, "COL_EXCEL": v} for k, v in dest_to_col.items()]
                st.dataframe(pd.DataFrame(m), width="stretch", hide_index=True)

            st.markdown("**Vista previa (primeras 50 filas)**")
            prev_df = pd.DataFrame(rows_to_add).fillna("")
            st.dataframe(prev_df.head(50), width="stretch", hide_index=True)
            st.caption(f"Filas detectadas para agregar: {len(rows_to_add)}")

            confirm_add = st.checkbox("Confirmo que deseo agregar estas filas al final", key="confirm_add_xlsx_sup2400205")

            if st.button("📥 Agregar filas al Sheet", key="btn_sup2400205_add_xlsx"):
                if not confirm_add:
                    st.warning("Activa la confirmación antes de agregar.")
                    st.stop()

                if str(periodo_up).strip() == "":
                    st.error("PERIODO es obligatorio.")
                    st.stop()

                payload = {
                    "action": "BATCH_ADD",
                    "sheet": SHEET_NAME_SUP,   # "SUP2400205"
                    "rows": rows_to_add
                }

                resp = requests.post(APPS_SCRIPT_URL, json=payload, timeout=120)
                data = resp.json() if resp.ok else {"ok": False, "error": resp.text}

                if data.get("ok"):
                    st.success(
                        f"Agregado ✅ Filas: {data.get('added')} | "
                        f"ID_SUP: {data.get('id_first')}–{data.get('id_last')}"
                    )
                    st.rerun()
                else:
                    st.error(data)

        except Exception as ex:
            st.error(f"No se pudo procesar el Excel: {ex}")

    # =========================
    # TABLA EDITABLE
    # =========================
    st.divider()
    st.subheader("TABLA EDITABLE")

    editable_cols_sup = [c for c in dfsup.columns if c != "ID_SUP"]
    viewsup = dfsup_f[["ID_SUP"] + editable_cols_sup].copy()

    orig_key_sup = f"orig_sup2400205_{f_anio}_{f_mes}_{f_periodo}_{f_esp}_{f_supervisor}_{f_apoyo}"
    if orig_key_sup not in st.session_state:
        st.session_state[orig_key_sup] = viewsup.copy()

    def _norm_sup(x):
        if x is None:
            return ""
        s = str(x)
        return "" if s.lower() in ("nan", "none", "null") else s.strip()

    editedsup = st.data_editor(
        viewsup,
        width="stretch",
        hide_index=True,
        num_rows="fixed",
        key=f"editor_sup2400205_{f_anio}_{f_mes}_{f_periodo}_{f_esp}_{f_supervisor}_{f_apoyo}",
    )

    if st.button("💾 Guardar SUP2400205", key="save_sup2400205"):
        original = st.session_state[orig_key_sup].copy()

        changed = pd.Series(False, index=editedsup.index)
        for c in editable_cols_sup:
            changed |= editedsup[c].map(_norm_sup) != original[c].map(_norm_sup)

        changed_rows = editedsup.loc[changed].copy()

        if changed_rows.empty:
            st.info("No hay cambios.")
        else:
            updates = []
            for _, r in changed_rows.iterrows():
                fields = {c: _norm_sup(r[c]) for c in editable_cols_sup}
                updates.append({"ID": _norm_sup(r["ID_SUP"]), "FIELDS": fields})

            payload = {"action": "BATCH_UPDATE", "sheet": SHEET_NAME_SUP, "updates": updates}
            resp = requests.post(APPS_SCRIPT_URL, json=payload, timeout=30)
            data = resp.json() if resp.ok else {"ok": False, "error": resp.text}

            if data.get("ok"):
                st.success(f"Guardado ✅ Filas actualizadas: {data.get('updated')} | No encontradas: {data.get('notFound')}")
                st.session_state[orig_key_sup] = editedsup.copy()
                st.rerun()
            else:
                st.error(data)

    # =========================
    # BORRAR FILA(S) + VISTA PREVIA (1 / Rango / Lista)
    # =========================
    st.divider()
    st.subheader("BORRAR FILA(S)")

    ids_sup = dfsup["ID_SUP"].dropna().astype(str).str.strip()
    ids_sup = ids_sup[ids_sup != ""].tolist()

    ids_int = [(_try_int(i), i) for i in ids_sup]
    if ids_int and all(v[0] is not None for v in ids_int):
        ids_sup_sorted = [v[1] for v in sorted(ids_int, key=lambda t: t[0])]
    else:
        ids_sup_sorted = sorted(ids_sup)

    modo = st.radio(
        "Modo de borrado",
        ["Uno", "Rango", "Lista"],
        horizontal=True,
        key="del_mode_sup2400205",
    )

    confirm = st.checkbox("Confirmo borrado irreversible", key="confirm_del_sup2400205")

    def preview_ids(target_ids):
        target_ids = [str(x).strip() for x in target_ids if str(x).strip() != ""]
        if not target_ids:
            return dfsup.head(0)

        dfprev = dfsup.copy()
        dfprev["ID_SUP"] = dfprev["ID_SUP"].astype(str).str.strip()
        dfprev = dfprev[dfprev["ID_SUP"].isin(target_ids)]

        tmp = dfprev["ID_SUP"].map(_try_int)
        if tmp.notna().all() and len(tmp) > 0:
            dfprev = dfprev.assign(_ord=tmp).sort_values("_ord").drop(columns=["_ord"])

        cols = ["ID_SUP"] + [c for c in dfprev.columns if c != "ID_SUP"]
        return dfprev[cols]

    def delete_one(id_value: str):
        payload = {"action": "DELETE", "sheet": SHEET_NAME_SUP, "id": str(id_value)}
        resp = requests.post(APPS_SCRIPT_URL, json=payload, timeout=30)
        if not resp.ok:
            return False, resp.text
        data = resp.json()
        return bool(data.get("ok")), data

    if modo == "Uno":
        id_del = st.selectbox("ID_SUP a borrar", ids_sup_sorted, key="del_one_sup2400205")

        st.markdown("**Vista previa**")
        prev = preview_ids([id_del])
        if prev.empty:
            st.info("No se encontró la fila para ese ID_SUP.")
        else:
            st.dataframe(prev, width="stretch", hide_index=True)

        if st.button("🗑️ Borrar 1", key="btn_del_one_2400205"):
            if not confirm:
                st.warning("Confirma antes de borrar.")
            else:
                ok, data = delete_one(id_del)
                if ok:
                    st.success(f"Borrado ✅ ID_SUP={id_del}")
                    st.rerun()
                else:
                    st.error(data)

    elif modo == "Rango":
        colr1, colr2 = st.columns(2)
        with colr1:
            desde = st.text_input("Desde ID_SUP", key="del_range_from_sup2400205")
        with colr2:
            hasta = st.text_input("Hasta ID_SUP", key="del_range_to_sup2400205")

        st.caption("Ejemplo: desde 10 hasta 25 (incluye ambos).")

        a = _try_int(desde)
        b = _try_int(hasta)

        if desde.strip() != "" and hasta.strip() != "" and (a is None or b is None):
            st.error("Para rango, 'Desde' y 'Hasta' deben ser números (ID_SUP numérico).")

        if a is not None and b is not None:
            if a > b:
                a, b = b, a

            target_ids = [str(i) for i in range(a, b + 1)]

            st.markdown("**Vista previa**")
            prev = preview_ids(target_ids)
            if prev.empty:
                st.info("No se encontraron filas dentro del rango en la hoja.")
            else:
                st.dataframe(prev, width="stretch", hide_index=True)
                st.caption(f"Filas encontradas: {len(prev)} (IDs inexistentes se ignorarán).")

            if st.button("🗑️ Borrar rango", key="btn_del_range_sup2400205"):
                if not confirm:
                    st.warning("Confirma antes de borrar.")
                else:
                    borrados = 0
                    fallos = []
                    for tid in target_ids:
                        ok, data = delete_one(tid)
                        if ok:
                            borrados += 1
                        else:
                            fallos.append((tid, data))

                    if borrados:
                        st.success(f"Borrado ✅ {borrados} filas.")
                    if fallos:
                        st.warning(f"No se pudieron borrar {len(fallos)} IDs.")
                        st.write(fallos)

                    st.rerun()

    else:  # Lista
        ids_texto = st.text_area(
            "IDs a borrar (separados por coma o salto de línea)",
            placeholder="Ej: 1,3,7,20\n25\n30",
            key="del_list_sup2400205",
        )

        raw = ids_texto.replace("\n", ",")
        parts = [p.strip() for p in raw.split(",") if p.strip() != ""]

        seen = set()
        target_ids = []
        for p in parts:
            if p not in seen:
                seen.add(p)
                target_ids.append(p)

        st.markdown("**Vista previa**")
        prev = preview_ids(target_ids)
        if not target_ids:
            st.info("Ingresa IDs para ver vista previa.")
        elif prev.empty:
            st.info("No se encontraron filas con esos IDs en la hoja.")
        else:
            st.dataframe(prev, width="stretch", hide_index=True)
            st.caption(f"Filas encontradas: {len(prev)} (IDs inexistentes se ignorarán).")

        if st.button("🗑️ Borrar lista", key="btn_del_list_sup2400205"):
            if not confirm:
                st.warning("Confirma antes de borrar.")
            else:
                if not target_ids:
                    st.error("Ingresa al menos un ID.")
                    st.stop()

                borrados = 0
                fallos = []
                for tid in target_ids:
                    ok, data = delete_one(tid)
                    if ok:
                        borrados += 1
                    else:
                        fallos.append((tid, data))

                if borrados:
                    st.success(f"Borrado ✅ {borrados} filas.")
                if fallos:
                    st.warning(f"No se pudieron borrar {len(fallos)} IDs.")
                    st.write(fallos)

                st.rerun()
                
# =========================
# TAB: SUP2500029
# =========================
with tab_SUP2500029:
    st.header("Contrato SUP2500029")

    SHEET_NAME_SUP = "SUP2500029"

    # =========================
    # LEER GOOGLE SHEET (CSV público)
    # =========================
    sheet_sup = urllib.parse.quote(SHEET_NAME_SUP)
    url_sup = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/gviz/tq?tqx=out:csv&sheet={sheet_sup}"

    dfsup = pd.read_csv(url_sup, dtype=str)
    dfsup.columns = (
        dfsup.columns.astype(str)
        .str.replace("\ufeff", "", regex=False)
        .str.strip()
        .str.upper()
    )
    dfsup = dfsup.loc[:, ~dfsup.columns.str.startswith("UNNAMED")]

    if "ID_SUP" not in dfsup.columns:
        st.error("La hoja 'SUP2500029' debe tener una columna 'ID_SUP' (clave única).")
        st.stop()

    dfsup["ID_SUP"] = dfsup["ID_SUP"].astype(str).str.strip()
    dfsup = dfsup[dfsup["ID_SUP"] != ""]
    dfsup = dfsup.fillna("")

    # =========================
    # HELPERS
    # =========================
    def pick_col(df, candidates):
        cols = set(df.columns)
        for c in candidates:
            if c in cols:
                return c
        return None

    def _try_int(x):
        try:
            return int(str(x).strip())
        except:
            return None

    def clean_opts(s: pd.Series):
        s = s.dropna().astype(str).str.strip().replace("", np.nan).dropna()
        nums = pd.to_numeric(s, errors="coerce")
        if nums.notna().all() and len(nums) > 0:
            order = sorted(nums.unique().tolist())
            out = []
            for v in order:
                if float(v).is_integer():
                    out.append(str(int(v)))
                else:
                    out.append(str(v))
            return out
        return sorted(s.unique().tolist())

    def _eq(df, col, val):
        return df[col].astype(str).str.strip() == str(val).strip()

    # =========================
    # FILTRO (Año, Mes, Periodo, Especialista, Supervisor, Apoyo)
    # =========================
    st.subheader("FILTRO")

    col_anio = pick_col(dfsup, ["AÑO", "ANIO", "ANO", "YEAR"])
    col_mes = pick_col(dfsup, ["MES", "MONTH"])
    col_periodo = pick_col(dfsup, ["PERIODO", "PERÍODO", "PERIODO.", "PERÍODO."])
    col_esp = pick_col(dfsup, ["ESPECIALISTA"])
    col_sup = pick_col(dfsup, ["SUPERVISOR", "NOMBRE COMPLETO DE SUPERVISOR", "NOMBRE COMPLETO DEL SUPERVISOR"])
    col_apoyo = pick_col(dfsup, ["APOYO"])

    opts_anio = ["TODOS"] + (clean_opts(dfsup[col_anio]) if col_anio else [])
    opts_mes = ["TODOS"] + (clean_opts(dfsup[col_mes]) if col_mes else [])
    opts_periodo = ["TODOS"] + (clean_opts(dfsup[col_periodo]) if col_periodo else [])
    opts_esp = ["TODOS"] + (clean_opts(dfsup[col_esp]) if col_esp else [])
    opts_sup = ["TODOS"] + (clean_opts(dfsup[col_sup]) if col_sup else [])
    opts_apoyo = ["TODOS"] + (clean_opts(dfsup[col_apoyo]) if col_apoyo else [])

    c1, c2, c3 = st.columns(3)
    c4, c5, c6 = st.columns(3)

    with c1:
        f_anio = st.selectbox("Año", opts_anio, key="f_sup2500029_anio", disabled=(col_anio is None))
    with c2:
        f_mes = st.selectbox("Mes", opts_mes, key="f_sup2500029_mes", disabled=(col_mes is None))
    with c3:
        f_periodo = st.selectbox("Periodo", opts_periodo, key="f_sup2500029_periodo", disabled=(col_periodo is None))

    with c4:
        f_esp = st.selectbox("Especialista", opts_esp, key="f_sup2500029_esp", disabled=(col_esp is None))
    with c5:
        f_supervisor = st.selectbox("Supervisor", opts_sup, key="f_sup2500029_supervisor", disabled=(col_sup is None))
    with c6:
        f_apoyo = st.selectbox("Apoyo", opts_apoyo, key="f_sup2500029_apoyo", disabled=(col_apoyo is None))

    dfsup_f = dfsup.copy()
    if col_anio and f_anio != "TODOS":
        dfsup_f = dfsup_f[_eq(dfsup_f, col_anio, f_anio)]
    if col_mes and f_mes != "TODOS":
        dfsup_f = dfsup_f[_eq(dfsup_f, col_mes, f_mes)]
    if col_periodo and f_periodo != "TODOS":
        dfsup_f = dfsup_f[_eq(dfsup_f, col_periodo, f_periodo)]
    if col_esp and f_esp != "TODOS":
        dfsup_f = dfsup_f[_eq(dfsup_f, col_esp, f_esp)]
    if col_sup and f_supervisor != "TODOS":
        dfsup_f = dfsup_f[_eq(dfsup_f, col_sup, f_supervisor)]
    if col_apoyo and f_apoyo != "TODOS":
        dfsup_f = dfsup_f[_eq(dfsup_f, col_apoyo, f_apoyo)]
    
    # =========================
    # CARGA DESDE EXCEL (XLSX) -> AGREGA FILAS A SUP2300128
    # =========================
    st.divider()
    st.subheader("CARGAR EXCEL (.xlsx) Y AGREGAR FILAS")

    # INPUTS (meta)
    anio_up = st.selectbox("AÑO", list(range(2025, 2036)), index=0, key="up_sup2500029_anio")
    mes_up = st.selectbox(
        "MES",
        ["ENERO","FEBRERO","MARZO","ABRIL","MAYO","JUNIO","JULIO","AGOSTO","SEPTIEMBRE","OCTUBRE","NOVIEMBRE","DICIEMBRE"],
        key="up_sup2500029_mes",
    )
    periodo_up = st.text_input("PERIODO", key="up_sup2500029_periodo")

    CONTRATO_AUTO = "SUP2500029"
    EST_AUTO = "VASMOL S.A.C."

    st.text_input("CONTRATO", value=CONTRATO_AUTO, disabled=False, key="up_sup2500029_contrato")
    st.text_input("EST", value=EST_AUTO, disabled=False, key="up_sup2500029_est")

    uploaded = st.file_uploader("Sube el Excel (.xlsx)", type=["xlsx"], key="up_sup2500029_file")
    
    HEADER_ROW_MAIN = 15
    HEADER_ROW_FALLBACK = 14
    DATA_START_ROW = 16

    # DESTINO = headers del SHEET (SUP2500029) | ORIGEN = headers del Excel subido
    MAP_DEST_TO_SRC_ALTS = {
        "ALCANCE SEGÚN CONTRATO": ["Alcance según contrato"],
        "PROCEDIMIENTO / TEMA: DETALLE": ["Procedimiento / Tema", "Procedimiento / Tema: Detalle", "Detalle"],
        "PROCEDIMIENTO / TEMA: ACTIVIDAD A REALIZAR": ["Actividad a realizar", "Procedimiento / Tema: Actividad a realizar"],
        "UNIDAD OPERATIVA A SUPERVISAR: AGENTE SUPERVISADO": [
            "Unidad operativa a Supervisar / Fiscalizar: Empresa",
            "Unidad operativa a Supervisar: Agente Supervisado",
            "Agente Supervisado",
            "Empresa",
        ],
    
        # Opcionales:
        "UNIDAD OPERATIVA A SUPERVISAR: UNIDAD /EXPEDIENTE": ["Unidad operativa a Supervisar: Unidad /Expediente", "Unidad /Expediente", "Unidad / Expediente"],
        "UNIDAD OPERATIVA A SUPERVISAR: UBIGEO": ["Unidad operativa a Supervisar: Ubigeo", "Ubigeo"],
        "FECHA: EJECUCIÓN": ["Fecha: Ejecución", "Ejecución"],
        "FECHA: ENTREGABLE": ["Entrega de Informe", "Fecha: Entregable", "Entregable"],
        "ENTREGABLES": ["Observaciones", "Entregables", "Entregable"],
        "ESPECIALISTA": ["Especialista"],
        "SUPERVISOR": ["Supervisor"],
        "APOYO": ["Apoyo"], 
    }

    
    OPTIONAL_DESTS = {
        "UNIDAD OPERATIVA A SUPERVISAR: UNIDAD /EXPEDIENTE",
        "UNIDAD OPERATIVA A SUPERVISAR: UBIGEO",
        "FECHA: EJECUCIÓN",
        "ESPECIALISTA",
        "SUPERVISOR",
        "APOYO",
    }
    
    REQUIRED_DESTS = [k for k in MAP_DEST_TO_SRC_ALTS.keys() if k not in OPTIONAL_DESTS]
    
    META = {
        "AÑO": lambda: safe_str(anio_up),
        "MES": lambda: safe_str(mes_up),
        "PERIODO": lambda: safe_str(periodo_up),
        "CONTRATO": lambda: CONTRATO_AUTO,
        "EST": lambda: EST_AUTO,
    }
    
    def excel_to_rows(file_bytes: bytes):
        wb = load_workbook(BytesIO(file_bytes), data_only=True)

        best_ws = None
        best_found = None
        best_score = -1
        for ws in wb.worksheets:
            found = build_found(ws)
            sc = score_sheet(found)
            if sc > best_score:
                best_score = sc
                best_ws = ws
                best_found = found

        if best_ws is None or best_score < max(4, len(REQUIRED_DESTS) // 2):
            raise ValueError("No pude identificar la hoja correcta (encabezados no coinciden en ninguna hoja).")

        dest_to_col, missing_required = resolve_columns(best_found)
        if missing_required:
            msg = "Faltan encabezados requeridos en el Excel subido:\n"
            for dest, alts in missing_required:
                msg += f"- {dest} (busqué: {alts})\n"
            raise ValueError(msg)

        diag = []
        for col in range(1, best_ws.max_column + 1):
            ht = header_text(best_ws, col)
            if ht:
                diag.append({"COL": col, "HEADER_DETECTADO": ht})
        diag_df = pd.DataFrame(diag)

        data = []
        r = DATA_START_ROW
        while True:
            row_vals = {}
            for dest, col in dest_to_col.items():
                row_vals[dest] = "" if col is None else best_ws.cell(r, col).value
            if row_is_empty(list(row_vals.values())):
                break
            data.append(row_vals)
            r += 1

        if not data:
            raise ValueError("No se detectaron filas de datos (desde la fila 16).")

        # Regla A
        k_det = "PROCEDIMIENTO / TEMA: DETALLE"
        k_act = "PROCEDIMIENTO / TEMA: ACTIVIDAD A REALIZAR"
        k_eje = "FECHA: EJECUCIÓN"
        k_ent1 = "FECHA: ENTREGABLE"
        k_ent2 = "ENTREGABLES"
        k_alc = "ALCANCE SEGÚN CONTRATO"

        for i in range(1, len(data)):
            row = data[i]
            prev = data[i - 1]
            if is_blank(row[k_det]) and is_blank(row[k_eje]) and is_blank(row[k_ent1]) and is_blank(row[k_ent2]):
                row[k_det] = prev[k_det]
                row[k_eje] = prev[k_eje]
                row[k_ent1] = prev[k_ent1]
                row[k_ent2] = prev[k_ent2]

        # Regla B
        ffill_keys = [k_alc, k_det, k_act]
        last_vals = {k: None for k in ffill_keys}
        for row in data:
            for k in ffill_keys:
                if is_blank(row[k]):
                    row[k] = last_vals[k]
                else:
                    last_vals[k] = row[k]

        rows = []
        for row in data:
            out = {}
            for dest in MAP_DEST_TO_SRC_ALTS.keys():
                out[dest] = safe_str(row.get(dest, ""))
            for k, fn in META.items():
                out[k] = safe_str(fn())
            rows.append(out)

        return rows, diag_df, best_ws.title, dest_to_col

    if uploaded:
        try:
            rows_to_add, diag_df, sheet_used, dest_to_col = excel_to_rows(uploaded.getvalue())

            st.info(f"Hoja detectada en el Excel: **{sheet_used}**")

            with st.expander("🔍 Encabezados detectados (fila 14/15 combinadas)"):
                st.dataframe(diag_df, width="stretch", hide_index=True)

            with st.expander("🧭 Mapeo usado (DESTINO en Sheet → columna Excel)"):
                m = [{"DESTINO (SHEET)": k, "COL_EXCEL": v} for k, v in dest_to_col.items()]
                st.dataframe(pd.DataFrame(m), width="stretch", hide_index=True)

            st.markdown("**Vista previa (primeras 50 filas)**")
            prev_df = pd.DataFrame(rows_to_add).fillna("")
            st.dataframe(prev_df.head(50), width="stretch", hide_index=True)
            st.caption(f"Filas detectadas para agregar: {len(rows_to_add)}")

            confirm_add = st.checkbox("Confirmo que deseo agregar estas filas al final", key="confirm_add_xlsx_sup2500029")

            if st.button("📥 Agregar filas al Sheet", key="btn_sup2500029_add_xlsx"):
                if not confirm_add:
                    st.warning("Activa la confirmación antes de agregar.")
                    st.stop()

                if str(periodo_up).strip() == "":
                    st.error("PERIODO es obligatorio.")
                    st.stop()

                payload = {
                    "action": "BATCH_ADD",
                    "sheet": SHEET_NAME_SUP,   # "SUP2500029"
                    "rows": rows_to_add
                }

                resp = requests.post(APPS_SCRIPT_URL, json=payload, timeout=120)
                data = resp.json() if resp.ok else {"ok": False, "error": resp.text}

                if data.get("ok"):
                    st.success(
                        f"Agregado ✅ Filas: {data.get('added')} | "
                        f"ID_SUP: {data.get('id_first')}–{data.get('id_last')}"
                    )
                    st.rerun()
                else:
                    st.error(data)

        except Exception as ex:
            st.error(f"No se pudo procesar el Excel: {ex}")

    # =========================
    # TABLA EDITABLE
    # =========================
    st.divider()
    st.subheader("TABLA EDITABLE")

    editable_cols_sup = [c for c in dfsup.columns if c != "ID_SUP"]
    viewsup = dfsup_f[["ID_SUP"] + editable_cols_sup].copy()

    orig_key_sup = f"orig_sup2500029_{f_anio}_{f_mes}_{f_periodo}_{f_esp}_{f_supervisor}_{f_apoyo}"
    if orig_key_sup not in st.session_state:
        st.session_state[orig_key_sup] = viewsup.copy()

    def _norm_sup(x):
        if x is None:
            return ""
        s = str(x)
        return "" if s.lower() in ("nan", "none", "null") else s.strip()

    editedsup = st.data_editor(
        viewsup,
        width="stretch",
        hide_index=True,
        num_rows="fixed",
        key=f"editor_sup2500029_{f_anio}_{f_mes}_{f_periodo}_{f_esp}_{f_supervisor}_{f_apoyo}",
    )

    if st.button("💾 Guardar SUP2500029", key="save_sup2500029"):
        original = st.session_state[orig_key_sup].copy()

        changed = pd.Series(False, index=editedsup.index)
        for c in editable_cols_sup:
            changed |= editedsup[c].map(_norm_sup) != original[c].map(_norm_sup)

        changed_rows = editedsup.loc[changed].copy()

        if changed_rows.empty:
            st.info("No hay cambios.")
        else:
            updates = []
            for _, r in changed_rows.iterrows():
                fields = {c: _norm_sup(r[c]) for c in editable_cols_sup}
                updates.append({"ID": _norm_sup(r["ID_SUP"]), "FIELDS": fields})

            payload = {"action": "BATCH_UPDATE", "sheet": SHEET_NAME_SUP, "updates": updates}
            resp = requests.post(APPS_SCRIPT_URL, json=payload, timeout=30)
            data = resp.json() if resp.ok else {"ok": False, "error": resp.text}

            if data.get("ok"):
                st.success(f"Guardado ✅ Filas actualizadas: {data.get('updated')} | No encontradas: {data.get('notFound')}")
                st.session_state[orig_key_sup] = editedsup.copy()
                st.rerun()
            else:
                st.error(data)

    # =========================
    # BORRAR FILA(S) + VISTA PREVIA (1 / Rango / Lista)
    # =========================
    st.divider()
    st.subheader("BORRAR FILA(S)")

    ids_sup = dfsup["ID_SUP"].dropna().astype(str).str.strip()
    ids_sup = ids_sup[ids_sup != ""].tolist()

    ids_int = [(_try_int(i), i) for i in ids_sup]
    if ids_int and all(v[0] is not None for v in ids_int):
        ids_sup_sorted = [v[1] for v in sorted(ids_int, key=lambda t: t[0])]
    else:
        ids_sup_sorted = sorted(ids_sup)

    modo = st.radio(
        "Modo de borrado",
        ["Uno", "Rango", "Lista"],
        horizontal=True,
        key="del_mode_sup2500029",
    )

    confirm = st.checkbox("Confirmo borrado irreversible", key="confirm_del_sup2500029")

    def preview_ids(target_ids):
        target_ids = [str(x).strip() for x in target_ids if str(x).strip() != ""]
        if not target_ids:
            return dfsup.head(0)

        dfprev = dfsup.copy()
        dfprev["ID_SUP"] = dfprev["ID_SUP"].astype(str).str.strip()
        dfprev = dfprev[dfprev["ID_SUP"].isin(target_ids)]

        tmp = dfprev["ID_SUP"].map(_try_int)
        if tmp.notna().all() and len(tmp) > 0:
            dfprev = dfprev.assign(_ord=tmp).sort_values("_ord").drop(columns=["_ord"])

        cols = ["ID_SUP"] + [c for c in dfprev.columns if c != "ID_SUP"]
        return dfprev[cols]

    def delete_one(id_value: str):
        payload = {"action": "DELETE", "sheet": SHEET_NAME_SUP, "id": str(id_value)}
        resp = requests.post(APPS_SCRIPT_URL, json=payload, timeout=30)
        if not resp.ok:
            return False, resp.text
        data = resp.json()
        return bool(data.get("ok")), data

    if modo == "Uno":
        id_del = st.selectbox("ID_SUP a borrar", ids_sup_sorted, key="del_one_sup2500029")

        st.markdown("**Vista previa**")
        prev = preview_ids([id_del])
        if prev.empty:
            st.info("No se encontró la fila para ese ID_SUP.")
        else:
            st.dataframe(prev, width="stretch", hide_index=True)

        if st.button("🗑️ Borrar 1", key="btn_del_one_2500029"):
            if not confirm:
                st.warning("Confirma antes de borrar.")
            else:
                ok, data = delete_one(id_del)
                if ok:
                    st.success(f"Borrado ✅ ID_SUP={id_del}")
                    st.rerun()
                else:
                    st.error(data)

    elif modo == "Rango":
        colr1, colr2 = st.columns(2)
        with colr1:
            desde = st.text_input("Desde ID_SUP", key="del_range_from_sup2500029")
        with colr2:
            hasta = st.text_input("Hasta ID_SUP", key="del_range_to_sup2500029")

        st.caption("Ejemplo: desde 10 hasta 25 (incluye ambos).")

        a = _try_int(desde)
        b = _try_int(hasta)

        if desde.strip() != "" and hasta.strip() != "" and (a is None or b is None):
            st.error("Para rango, 'Desde' y 'Hasta' deben ser números (ID_SUP numérico).")

        if a is not None and b is not None:
            if a > b:
                a, b = b, a

            target_ids = [str(i) for i in range(a, b + 1)]

            st.markdown("**Vista previa**")
            prev = preview_ids(target_ids)
            if prev.empty:
                st.info("No se encontraron filas dentro del rango en la hoja.")
            else:
                st.dataframe(prev, width="stretch", hide_index=True)
                st.caption(f"Filas encontradas: {len(prev)} (IDs inexistentes se ignorarán).")

            if st.button("🗑️ Borrar rango", key="btn_del_range_sup2500029"):
                if not confirm:
                    st.warning("Confirma antes de borrar.")
                else:
                    borrados = 0
                    fallos = []
                    for tid in target_ids:
                        ok, data = delete_one(tid)
                        if ok:
                            borrados += 1
                        else:
                            fallos.append((tid, data))

                    if borrados:
                        st.success(f"Borrado ✅ {borrados} filas.")
                    if fallos:
                        st.warning(f"No se pudieron borrar {len(fallos)} IDs.")
                        st.write(fallos)

                    st.rerun()

    else:  # Lista
        ids_texto = st.text_area(
            "IDs a borrar (separados por coma o salto de línea)",
            placeholder="Ej: 1,3,7,20\n25\n30",
            key="del_list_sup2500029",
        )

        raw = ids_texto.replace("\n", ",")
        parts = [p.strip() for p in raw.split(",") if p.strip() != ""]

        seen = set()
        target_ids = []
        for p in parts:
            if p not in seen:
                seen.add(p)
                target_ids.append(p)

        st.markdown("**Vista previa**")
        prev = preview_ids(target_ids)
        if not target_ids:
            st.info("Ingresa IDs para ver vista previa.")
        elif prev.empty:
            st.info("No se encontraron filas con esos IDs en la hoja.")
        else:
            st.dataframe(prev, width="stretch", hide_index=True)
            st.caption(f"Filas encontradas: {len(prev)} (IDs inexistentes se ignorarán).")

        if st.button("🗑️ Borrar lista", key="btn_del_list_sup2500029"):
            if not confirm:
                st.warning("Confirma antes de borrar.")
            else:
                if not target_ids:
                    st.error("Ingresa al menos un ID.")
                    st.stop()

                borrados = 0
                fallos = []
                for tid in target_ids:
                    ok, data = delete_one(tid)
                    if ok:
                        borrados += 1
                    else:
                        fallos.append((tid, data))

                if borrados:
                    st.success(f"Borrado ✅ {borrados} filas.")
                if fallos:
                    st.warning(f"No se pudieron borrar {len(fallos)} IDs.")
                    st.write(fallos)

                st.rerun()
                
# =========================
# TAB: SUP2400028
# =========================
with tab_SUP2400028:
    st.header("Contrato SUP2400028")

    SHEET_NAME_SUP = "SUP2400028"

    # =========================
    # LEER GOOGLE SHEET (CSV público)
    # =========================
    sheet_sup = urllib.parse.quote(SHEET_NAME_SUP)
    url_sup = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/gviz/tq?tqx=out:csv&sheet={sheet_sup}"

    dfsup = pd.read_csv(url_sup, dtype=str)
    dfsup.columns = (
        dfsup.columns.astype(str)
        .str.replace("\ufeff", "", regex=False)
        .str.strip()
        .str.upper()
    )
    dfsup = dfsup.loc[:, ~dfsup.columns.str.startswith("UNNAMED")]

    if "ID_SUP" not in dfsup.columns:
        st.error("La hoja 'SUP2400028' debe tener una columna 'ID_SUP' (clave única).")
        st.stop()

    dfsup["ID_SUP"] = dfsup["ID_SUP"].astype(str).str.strip()
    dfsup = dfsup[dfsup["ID_SUP"] != ""]
    dfsup = dfsup.fillna("")

    # =========================
    # HELPERS
    # =========================
    def pick_col(df, candidates):
        cols = set(df.columns)
        for c in candidates:
            if c in cols:
                return c
        return None

    def _try_int(x):
        try:
            return int(str(x).strip())
        except:
            return None

    def clean_opts(s: pd.Series):
        s = s.dropna().astype(str).str.strip().replace("", np.nan).dropna()
        nums = pd.to_numeric(s, errors="coerce")
        if nums.notna().all() and len(nums) > 0:
            order = sorted(nums.unique().tolist())
            out = []
            for v in order:
                if float(v).is_integer():
                    out.append(str(int(v)))
                else:
                    out.append(str(v))
            return out
        return sorted(s.unique().tolist())

    def _eq(df, col, val):
        return df[col].astype(str).str.strip() == str(val).strip()

    # =========================
    # FILTRO (Año, Mes, Periodo, Especialista, Supervisor, Apoyo)
    # =========================
    st.subheader("FILTRO")

    col_anio = pick_col(dfsup, ["AÑO", "ANIO", "ANO", "YEAR"])
    col_mes = pick_col(dfsup, ["MES", "MONTH"])
    col_periodo = pick_col(dfsup, ["PERIODO", "PERÍODO", "PERIODO.", "PERÍODO."])
    col_esp = pick_col(dfsup, ["ESPECIALISTA"])
    col_sup = pick_col(dfsup, ["SUPERVISOR", "NOMBRE COMPLETO DE SUPERVISOR", "NOMBRE COMPLETO DEL SUPERVISOR"])
    col_apoyo = pick_col(dfsup, ["APOYO"])

    opts_anio = ["TODOS"] + (clean_opts(dfsup[col_anio]) if col_anio else [])
    opts_mes = ["TODOS"] + (clean_opts(dfsup[col_mes]) if col_mes else [])
    opts_periodo = ["TODOS"] + (clean_opts(dfsup[col_periodo]) if col_periodo else [])
    opts_esp = ["TODOS"] + (clean_opts(dfsup[col_esp]) if col_esp else [])
    opts_sup = ["TODOS"] + (clean_opts(dfsup[col_sup]) if col_sup else [])
    opts_apoyo = ["TODOS"] + (clean_opts(dfsup[col_apoyo]) if col_apoyo else [])

    c1, c2, c3 = st.columns(3)
    c4, c5, c6 = st.columns(3)

    with c1:
        f_anio = st.selectbox("Año", opts_anio, key="f_sup2400028_anio", disabled=(col_anio is None))
    with c2:
        f_mes = st.selectbox("Mes", opts_mes, key="f_sup2400028_mes", disabled=(col_mes is None))
    with c3:
        f_periodo = st.selectbox("Periodo", opts_periodo, key="f_sup2400028_periodo", disabled=(col_periodo is None))

    with c4:
        f_esp = st.selectbox("Especialista", opts_esp, key="f_sup2400028_esp", disabled=(col_esp is None))
    with c5:
        f_supervisor = st.selectbox("Supervisor", opts_sup, key="f_sup2400028_supervisor", disabled=(col_sup is None))
    with c6:
        f_apoyo = st.selectbox("Apoyo", opts_apoyo, key="f_sup2400028_apoyo", disabled=(col_apoyo is None))

    dfsup_f = dfsup.copy()
    if col_anio and f_anio != "TODOS":
        dfsup_f = dfsup_f[_eq(dfsup_f, col_anio, f_anio)]
    if col_mes and f_mes != "TODOS":
        dfsup_f = dfsup_f[_eq(dfsup_f, col_mes, f_mes)]
    if col_periodo and f_periodo != "TODOS":
        dfsup_f = dfsup_f[_eq(dfsup_f, col_periodo, f_periodo)]
    if col_esp and f_esp != "TODOS":
        dfsup_f = dfsup_f[_eq(dfsup_f, col_esp, f_esp)]
    if col_sup and f_supervisor != "TODOS":
        dfsup_f = dfsup_f[_eq(dfsup_f, col_sup, f_supervisor)]
    if col_apoyo and f_apoyo != "TODOS":
        dfsup_f = dfsup_f[_eq(dfsup_f, col_apoyo, f_apoyo)]
    
    # =========================
    # CARGA DESDE EXCEL (XLSX) -> AGREGA FILAS A SUP2300128
    # =========================
    st.divider()
    st.subheader("CARGAR EXCEL (.xlsx) Y AGREGAR FILAS")

    # INPUTS (meta)
    anio_up = st.selectbox("AÑO", list(range(2025, 2036)), index=0, key="up_sup2400028_anio")
    mes_up = st.selectbox(
        "MES",
        ["ENERO","FEBRERO","MARZO","ABRIL","MAYO","JUNIO","JULIO","AGOSTO","SEPTIEMBRE","OCTUBRE","NOVIEMBRE","DICIEMBRE"],
        key="up_sup2400028_mes",
    )
    periodo_up = st.text_input("PERIODO", key="up_sup2400028_periodo")

    CONTRATO_AUTO = "SUP2400028"
    EST_AUTO = "SEOUL INSPECTION AND TESTING CO. LTD. SUCURSAL DEL PERU"

    st.text_input("CONTRATO", value=CONTRATO_AUTO, disabled=False, key="up_sup2400028_contrato")
    st.text_input("EST", value=EST_AUTO, disabled=False, key="up_sup2400028_est")

    uploaded = st.file_uploader("Sube el Excel (.xlsx)", type=["xlsx"], key="up_sup2400028_file")
    
    HEADER_ROW_MAIN = 16
    HEADER_ROW_FALLBACK = 15
    DATA_START_ROW = 17

    # DESTINO = headers del SHEET (SUP2400028) | ORIGEN = headers del Excel subido
    MAP_DEST_TO_SRC_ALTS = {
        "ALCANCE SEGÚN CONTRATO": ["Alcance según contrato"],
        "PROCEDIMIENTO / TEMA: DETALLE": ["Procedimiento / Tema", "Procedimiento / Tema: Detalle", "Detalle"],
        "PROCEDIMIENTO / TEMA: ACTIVIDAD A REALIZAR": ["Actividad a realizar", "Procedimiento / Tema: Actividad a realizar"],
        "UNIDAD OPERATIVA A SUPERVISAR: AGENTE SUPERVISADO": [
            "Unidad operativa a Supervisar / Fiscalizar: Empresa",
            "Unidad operativa a Supervisar: Agente Supervisado",
            "Agente Supervisado",
            "Empresa",
        ],
    
        # Opcionales:
        "UNIDAD OPERATIVA A SUPERVISAR: UNIDAD /EXPEDIENTE": ["Unidad operativa a Supervisar: Unidad /Expediente", "Unidad /Expediente", "Unidad / Expediente"],
        "UNIDAD OPERATIVA A SUPERVISAR: UBIGEO": ["Unidad operativa a Supervisar: Ubigeo", "Ubigeo"],
        "FECHA: EJECUCIÓN": ["Fecha: Ejecución", "Ejecución"],
        "FECHA: ENTREGABLE": ["Entrega de Informe", "Fecha: Entregable", "Entregable"],
        "ENTREGABLES": ["Observaciones", "Entregables", "Entregable"],
        "ESPECIALISTA": ["Especialista"],
        "SUPERVISOR": ["Supervisor"],
        "APOYO": ["Apoyo"], 
    }

    
    OPTIONAL_DESTS = {
        "UNIDAD OPERATIVA A SUPERVISAR: UNIDAD /EXPEDIENTE",
        "UNIDAD OPERATIVA A SUPERVISAR: UBIGEO",
        "FECHA: EJECUCIÓN",
        "ESPECIALISTA",
        "SUPERVISOR",
        "APOYO",
    }
    
    REQUIRED_DESTS = [k for k in MAP_DEST_TO_SRC_ALTS.keys() if k not in OPTIONAL_DESTS]
    
    META = {
        "AÑO": lambda: safe_str(anio_up),
        "MES": lambda: safe_str(mes_up),
        "PERIODO": lambda: safe_str(periodo_up),
        "CONTRATO": lambda: CONTRATO_AUTO,
        "EST": lambda: EST_AUTO,
    }
    
    def excel_to_rows(file_bytes: bytes):
        wb = load_workbook(BytesIO(file_bytes), data_only=True)

        best_ws = None
        best_found = None
        best_score = -1
        for ws in wb.worksheets:
            found = build_found(ws)
            sc = score_sheet(found)
            if sc > best_score:
                best_score = sc
                best_ws = ws
                best_found = found

        if best_ws is None or best_score < max(4, len(REQUIRED_DESTS) // 2):
            raise ValueError("No pude identificar la hoja correcta (encabezados no coinciden en ninguna hoja).")

        dest_to_col, missing_required = resolve_columns(best_found)
        if missing_required:
            msg = "Faltan encabezados requeridos en el Excel subido:\n"
            for dest, alts in missing_required:
                msg += f"- {dest} (busqué: {alts})\n"
            raise ValueError(msg)

        diag = []
        for col in range(1, best_ws.max_column + 1):
            ht = header_text(best_ws, col)
            if ht:
                diag.append({"COL": col, "HEADER_DETECTADO": ht})
        diag_df = pd.DataFrame(diag)

        data = []
        r = DATA_START_ROW
        while True:
            row_vals = {}
            for dest, col in dest_to_col.items():
                row_vals[dest] = "" if col is None else best_ws.cell(r, col).value
            if row_is_empty(list(row_vals.values())):
                break
            data.append(row_vals)
            r += 1

        if not data:
            raise ValueError("No se detectaron filas de datos (desde la fila 17).")

        # Regla A
        k_det = "PROCEDIMIENTO / TEMA: DETALLE"
        k_act = "PROCEDIMIENTO / TEMA: ACTIVIDAD A REALIZAR"
        k_eje = "FECHA: EJECUCIÓN"
        k_ent1 = "FECHA: ENTREGABLE"
        k_ent2 = "ENTREGABLES"
        k_alc = "ALCANCE SEGÚN CONTRATO"

        for i in range(1, len(data)):
            row = data[i]
            prev = data[i - 1]
            if is_blank(row[k_det]) and is_blank(row[k_eje]) and is_blank(row[k_ent1]) and is_blank(row[k_ent2]):
                row[k_det] = prev[k_det]
                row[k_eje] = prev[k_eje]
                row[k_ent1] = prev[k_ent1]
                row[k_ent2] = prev[k_ent2]

        # Regla B
        ffill_keys = [k_alc, k_det, k_act]
        last_vals = {k: None for k in ffill_keys}
        for row in data:
            for k in ffill_keys:
                if is_blank(row[k]):
                    row[k] = last_vals[k]
                else:
                    last_vals[k] = row[k]

        rows = []
        for row in data:
            out = {}
            for dest in MAP_DEST_TO_SRC_ALTS.keys():
                out[dest] = safe_str(row.get(dest, ""))
            for k, fn in META.items():
                out[k] = safe_str(fn())
            rows.append(out)

        return rows, diag_df, best_ws.title, dest_to_col

    if uploaded:
        try:
            rows_to_add, diag_df, sheet_used, dest_to_col = excel_to_rows(uploaded.getvalue())

            st.info(f"Hoja detectada en el Excel: **{sheet_used}**")

            with st.expander("🔍 Encabezados detectados (fila 15/16 combinadas)"):
                st.dataframe(diag_df, width="stretch", hide_index=True)

            with st.expander("🧭 Mapeo usado (DESTINO en Sheet → columna Excel)"):
                m = [{"DESTINO (SHEET)": k, "COL_EXCEL": v} for k, v in dest_to_col.items()]
                st.dataframe(pd.DataFrame(m), width="stretch", hide_index=True)

            st.markdown("**Vista previa (primeras 50 filas)**")
            prev_df = pd.DataFrame(rows_to_add).fillna("")
            st.dataframe(prev_df.head(50), width="stretch", hide_index=True)
            st.caption(f"Filas detectadas para agregar: {len(rows_to_add)}")

            confirm_add = st.checkbox("Confirmo que deseo agregar estas filas al final", key="confirm_add_xlsx_sup2400028")

            if st.button("📥 Agregar filas al Sheet", key="btn_sup2400028_add_xlsx"):
                if not confirm_add:
                    st.warning("Activa la confirmación antes de agregar.")
                    st.stop()

                if str(periodo_up).strip() == "":
                    st.error("PERIODO es obligatorio.")
                    st.stop()

                payload = {
                    "action": "BATCH_ADD",
                    "sheet": SHEET_NAME_SUP,   # "SUP2400028"
                    "rows": rows_to_add
                }

                resp = requests.post(APPS_SCRIPT_URL, json=payload, timeout=120)
                data = resp.json() if resp.ok else {"ok": False, "error": resp.text}

                if data.get("ok"):
                    st.success(
                        f"Agregado ✅ Filas: {data.get('added')} | "
                        f"ID_SUP: {data.get('id_first')}–{data.get('id_last')}"
                    )
                    st.rerun()
                else:
                    st.error(data)

        except Exception as ex:
            st.error(f"No se pudo procesar el Excel: {ex}")

    # =========================
    # TABLA EDITABLE
    # =========================
    st.divider()
    st.subheader("TABLA EDITABLE")

    editable_cols_sup = [c for c in dfsup.columns if c != "ID_SUP"]
    viewsup = dfsup_f[["ID_SUP"] + editable_cols_sup].copy()

    orig_key_sup = f"orig_sup2400028_{f_anio}_{f_mes}_{f_periodo}_{f_esp}_{f_supervisor}_{f_apoyo}"
    if orig_key_sup not in st.session_state:
        st.session_state[orig_key_sup] = viewsup.copy()

    def _norm_sup(x):
        if x is None:
            return ""
        s = str(x)
        return "" if s.lower() in ("nan", "none", "null") else s.strip()

    editedsup = st.data_editor(
        viewsup,
        width="stretch",
        hide_index=True,
        num_rows="fixed",
        key=f"editor_sup2400028_{f_anio}_{f_mes}_{f_periodo}_{f_esp}_{f_supervisor}_{f_apoyo}",
    )

    if st.button("💾 Guardar SUP2400028", key="save_sup2400028"):
        original = st.session_state[orig_key_sup].copy()

        changed = pd.Series(False, index=editedsup.index)
        for c in editable_cols_sup:
            changed |= editedsup[c].map(_norm_sup) != original[c].map(_norm_sup)

        changed_rows = editedsup.loc[changed].copy()

        if changed_rows.empty:
            st.info("No hay cambios.")
        else:
            updates = []
            for _, r in changed_rows.iterrows():
                fields = {c: _norm_sup(r[c]) for c in editable_cols_sup}
                updates.append({"ID": _norm_sup(r["ID_SUP"]), "FIELDS": fields})

            payload = {"action": "BATCH_UPDATE", "sheet": SHEET_NAME_SUP, "updates": updates}
            resp = requests.post(APPS_SCRIPT_URL, json=payload, timeout=30)
            data = resp.json() if resp.ok else {"ok": False, "error": resp.text}

            if data.get("ok"):
                st.success(f"Guardado ✅ Filas actualizadas: {data.get('updated')} | No encontradas: {data.get('notFound')}")
                st.session_state[orig_key_sup] = editedsup.copy()
                st.rerun()
            else:
                st.error(data)

    # =========================
    # BORRAR FILA(S) + VISTA PREVIA (1 / Rango / Lista)
    # =========================
    st.divider()
    st.subheader("BORRAR FILA(S)")

    ids_sup = dfsup["ID_SUP"].dropna().astype(str).str.strip()
    ids_sup = ids_sup[ids_sup != ""].tolist()

    ids_int = [(_try_int(i), i) for i in ids_sup]
    if ids_int and all(v[0] is not None for v in ids_int):
        ids_sup_sorted = [v[1] for v in sorted(ids_int, key=lambda t: t[0])]
    else:
        ids_sup_sorted = sorted(ids_sup)

    modo = st.radio(
        "Modo de borrado",
        ["Uno", "Rango", "Lista"],
        horizontal=True,
        key="del_mode_sup2400028",
    )

    confirm = st.checkbox("Confirmo borrado irreversible", key="confirm_del_sup2400028")

    def preview_ids(target_ids):
        target_ids = [str(x).strip() for x in target_ids if str(x).strip() != ""]
        if not target_ids:
            return dfsup.head(0)

        dfprev = dfsup.copy()
        dfprev["ID_SUP"] = dfprev["ID_SUP"].astype(str).str.strip()
        dfprev = dfprev[dfprev["ID_SUP"].isin(target_ids)]

        tmp = dfprev["ID_SUP"].map(_try_int)
        if tmp.notna().all() and len(tmp) > 0:
            dfprev = dfprev.assign(_ord=tmp).sort_values("_ord").drop(columns=["_ord"])

        cols = ["ID_SUP"] + [c for c in dfprev.columns if c != "ID_SUP"]
        return dfprev[cols]

    def delete_one(id_value: str):
        payload = {"action": "DELETE", "sheet": SHEET_NAME_SUP, "id": str(id_value)}
        resp = requests.post(APPS_SCRIPT_URL, json=payload, timeout=30)
        if not resp.ok:
            return False, resp.text
        data = resp.json()
        return bool(data.get("ok")), data

    if modo == "Uno":
        id_del = st.selectbox("ID_SUP a borrar", ids_sup_sorted, key="del_one_sup2400028")

        st.markdown("**Vista previa**")
        prev = preview_ids([id_del])
        if prev.empty:
            st.info("No se encontró la fila para ese ID_SUP.")
        else:
            st.dataframe(prev, width="stretch", hide_index=True)

        if st.button("🗑️ Borrar 1", key="btn_del_one_2400028"):
            if not confirm:
                st.warning("Confirma antes de borrar.")
            else:
                ok, data = delete_one(id_del)
                if ok:
                    st.success(f"Borrado ✅ ID_SUP={id_del}")
                    st.rerun()
                else:
                    st.error(data)

    elif modo == "Rango":
        colr1, colr2 = st.columns(2)
        with colr1:
            desde = st.text_input("Desde ID_SUP", key="del_range_from_sup2400028")
        with colr2:
            hasta = st.text_input("Hasta ID_SUP", key="del_range_to_sup2400028")

        st.caption("Ejemplo: desde 10 hasta 25 (incluye ambos).")

        a = _try_int(desde)
        b = _try_int(hasta)

        if desde.strip() != "" and hasta.strip() != "" and (a is None or b is None):
            st.error("Para rango, 'Desde' y 'Hasta' deben ser números (ID_SUP numérico).")

        if a is not None and b is not None:
            if a > b:
                a, b = b, a

            target_ids = [str(i) for i in range(a, b + 1)]

            st.markdown("**Vista previa**")
            prev = preview_ids(target_ids)
            if prev.empty:
                st.info("No se encontraron filas dentro del rango en la hoja.")
            else:
                st.dataframe(prev, width="stretch", hide_index=True)
                st.caption(f"Filas encontradas: {len(prev)} (IDs inexistentes se ignorarán).")

            if st.button("🗑️ Borrar rango", key="btn_del_range_sup2400028"):
                if not confirm:
                    st.warning("Confirma antes de borrar.")
                else:
                    borrados = 0
                    fallos = []
                    for tid in target_ids:
                        ok, data = delete_one(tid)
                        if ok:
                            borrados += 1
                        else:
                            fallos.append((tid, data))

                    if borrados:
                        st.success(f"Borrado ✅ {borrados} filas.")
                    if fallos:
                        st.warning(f"No se pudieron borrar {len(fallos)} IDs.")
                        st.write(fallos)

                    st.rerun()

    else:  # Lista
        ids_texto = st.text_area(
            "IDs a borrar (separados por coma o salto de línea)",
            placeholder="Ej: 1,3,7,20\n25\n30",
            key="del_list_sup2400028",
        )

        raw = ids_texto.replace("\n", ",")
        parts = [p.strip() for p in raw.split(",") if p.strip() != ""]

        seen = set()
        target_ids = []
        for p in parts:
            if p not in seen:
                seen.add(p)
                target_ids.append(p)

        st.markdown("**Vista previa**")
        prev = preview_ids(target_ids)
        if not target_ids:
            st.info("Ingresa IDs para ver vista previa.")
        elif prev.empty:
            st.info("No se encontraron filas con esos IDs en la hoja.")
        else:
            st.dataframe(prev, width="stretch", hide_index=True)
            st.caption(f"Filas encontradas: {len(prev)} (IDs inexistentes se ignorarán).")

        if st.button("🗑️ Borrar lista", key="btn_del_list_sup2400028"):
            if not confirm:
                st.warning("Confirma antes de borrar.")
            else:
                if not target_ids:
                    st.error("Ingresa al menos un ID.")
                    st.stop()

                borrados = 0
                fallos = []
                for tid in target_ids:
                    ok, data = delete_one(tid)
                    if ok:
                        borrados += 1
                    else:
                        fallos.append((tid, data))

                if borrados:
                    st.success(f"Borrado ✅ {borrados} filas.")
                if fallos:
                    st.warning(f"No se pudieron borrar {len(fallos)} IDs.")
                    st.write(fallos)

                st.rerun()
                